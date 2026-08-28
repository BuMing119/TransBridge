from __future__ import annotations

from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook
import pytest

from tests.application.terminology.story08_support import build, decision, draft
from transbridge.application.terminology.in_memory import InMemoryTerminologyRepository
from transbridge.application.terminology.models import ArtifactKind, ArtifactStatus, DraftRef, TerminologyDraft
from transbridge.application.terminology.ports import PageRequest
from transbridge.application.terminology.renderers import (
    ArtifactPublishPolicy,
    ArtifactRenderCoordinator,
    ArtifactRenderError,
    pending_artifact,
)
from transbridge.application.terminology.renderers.quality_excel import QualityExcelRenderer
from transbridge.application.terminology.renderers.spreadsheet_safety import (
    EXCEL_CELL_CHAR_LIMIT,
    expanded_spreadsheet_rows,
    spreadsheet_chunks,
)
from transbridge.application.terminology.report_queries import TerminologyReportQueryService
from transbridge.application.terminology.reports import NoDraftIdentity, TerminologyReportSnapshotFactory
from transbridge.persistence.terminology import (
    SqliteTerminologyRepository,
    TerminologyConnectionFactory,
    TerminologyPaths,
)


def test_snapshot_freezes_build_and_full_pinned_draft_identity() -> None:
    repository = InMemoryTerminologyRepository()
    result = build()
    repository.put_build(result)
    original_draft = draft()
    snapshot = TerminologyReportSnapshotFactory(repository).freeze(result.ref, draft=original_draft)
    repository.put_report_snapshot(snapshot)

    changed_ref = DraftRef("draft-1", "project-1", "variant-1", None, "no-base", 1, "decision-set-1")
    changed_draft = TerminologyDraft(changed_ref, (decision(translation="巨龙"),))

    assert snapshot.draft_ref == original_draft.ref
    assert snapshot.terms == original_draft.decisions
    assert snapshot.terms != changed_draft.decisions
    assert repository.get_build(result.ref) is result


def test_snapshot_rejects_facts_that_contradict_the_pinned_draft() -> None:
    repository = InMemoryTerminologyRepository()
    result = build()
    repository.put_build(result)
    pinned = draft()
    injected = (decision("term-injected", translation="伪造"),)

    with pytest.raises(ValueError, match="terms must match"):
        TerminologyReportSnapshotFactory(repository).freeze(result.ref, draft=pinned, terms=injected)


def test_no_draft_sentinel_binds_base_digest_and_revision() -> None:
    repository = InMemoryTerminologyRepository()
    repository.put_build(build())
    factory = TerminologyReportSnapshotFactory(repository)

    first = factory.freeze(
        build().ref,
        no_draft=NoDraftIdentity("project-1", "variant-1", None, "no-base", revision=3),
    )
    second = factory.freeze(
        build().ref,
        no_draft=NoDraftIdentity("project-1", "variant-1", None, "changed-base", revision=3),
    )

    assert first.no_draft_identity is not None
    assert first.no_draft_identity != second.no_draft_identity
    assert first.ref != second.ref


def test_ui_pages_and_quality_excel_share_one_snapshot_ref_and_empty_workbook_has_four_tables(tmp_path: Path) -> None:
    repository = InMemoryTerminologyRepository()
    repository.put_build(build())
    snapshot = TerminologyReportSnapshotFactory(repository).freeze(
        build().ref,
        no_draft=NoDraftIdentity("project-1", "variant-1", None, "no-base"),
    )
    repository.put_report_snapshot(snapshot)
    queries = TerminologyReportQueryService(repository)

    page = queries.terms(snapshot.ref, PageRequest(limit=1))
    artifact = QualityExcelRenderer(queries).render(snapshot.ref, tmp_path / "quality.xlsx")
    workbook = load_workbook(artifact.path, read_only=True)

    assert page.snapshot_digest == snapshot.ref.content_digest
    assert artifact.semantic_manifest.fact_digest == page.snapshot_digest
    assert workbook.sheetnames == ["构建摘要", "术语对照", "同名异译", "人工调整记录"]
    assert tuple(cell.value for cell in next(workbook["术语对照"].iter_rows()))[:3] == ("术语ID", "原名", "译名")


def test_spreadsheet_values_escape_formulas_illegal_unicode_and_split_without_truncation() -> None:
    assert spreadsheet_chunks("=1+1") == ("'=1+1",)
    assert spreadsheet_chunks("bad\x00value") == ("bad\\u0000value",)

    original = "=" + "x" * (EXCEL_CELL_CHAR_LIMIT * 2)
    chunks = spreadsheet_chunks(original)

    assert len(chunks) == 3
    assert all(len(chunk) <= EXCEL_CELL_CHAR_LIMIT for chunk in chunks)
    assert "".join(chunk[1:] if chunk.startswith("'") else chunk for chunk in chunks) == original
    assert len(expanded_spreadsheet_rows(("id", original))) == 3


class _Ledger:
    def __init__(self) -> None:
        self.entries = {}

    def get_artifact(self, artifact_id):
        return self.entries.get(artifact_id)

    def put_artifact(self, entry):
        self.entries[entry.artifact_id] = entry
        return entry

    def update_artifact(self, entry, *, expected_status, expected_revision):
        assert self.entries[entry.artifact_id].status is expected_status
        assert self.entries[entry.artifact_id].revision == expected_revision
        return self.put_artifact(entry)


def test_renderer_failure_changes_only_ledger_state() -> None:
    repository = InMemoryTerminologyRepository()
    result = build()
    repository.put_build(result)
    ledger = _Ledger()
    entry = pending_artifact(
        owner_ref=result.ref.build_key,
        owner_digest=result.ref.content_digest,
        kind=ArtifactKind.QUALITY_EXCEL,
        renderer_version="test",
        target="quality.xlsx",
    )

    with pytest.raises(ArtifactRenderError, match="immutable terminology facts"):
        ArtifactRenderCoordinator(ledger).render(entry, lambda: (_ for _ in ()).throw(OSError("disk full")))

    assert ledger.entries[entry.artifact_id].status is ArtifactStatus.FAILED
    assert "disk full" in ledger.entries[entry.artifact_id].diagnostic
    assert repository.get_build(result.ref) is result


def test_default_publish_policy_refuses_overwrite_and_rename_is_explicit(tmp_path: Path) -> None:
    repository = InMemoryTerminologyRepository()
    repository.put_build(build())
    snapshot = TerminologyReportSnapshotFactory(repository).freeze(
        build().ref,
        no_draft=NoDraftIdentity("project-1", "variant-1", None, "no-base"),
    )
    repository.put_report_snapshot(snapshot)
    renderer = QualityExcelRenderer(TerminologyReportQueryService(repository))
    target = tmp_path / "quality.xlsx"
    first = renderer.render(snapshot.ref, target)

    with pytest.raises(FileExistsError):
        renderer.render(snapshot.ref, target)
    renamed = renderer.render(snapshot.ref, target, policy=ArtifactPublishPolicy.RENAME)

    assert first.path == target
    assert renamed.path.name == "quality-001.xlsx"


def test_quality_excel_splits_at_stable_row_boundary_without_losing_terms(tmp_path: Path) -> None:
    repository = InMemoryTerminologyRepository()
    repository.put_build(build())
    decisions = (
        replace(decision("term-1"), original="=Formula", normalized_original="=formula", notes="bad\x00note"),
        decision("term-2", translation="飞龙"),
        decision("term-3", translation="幼龙"),
    )
    source_draft = TerminologyDraft(draft().ref, decisions)
    snapshot = TerminologyReportSnapshotFactory(repository).freeze(build().ref, draft=source_draft)
    repository.put_report_snapshot(snapshot)

    artifact = QualityExcelRenderer(TerminologyReportQueryService(repository), max_sheet_rows=3).render(
        snapshot.ref,
        tmp_path / "split.xlsx",
    )
    workbook = load_workbook(artifact.path, read_only=True)
    ids: list[str] = []
    for sheet_name in ("术语对照", "术语对照-002"):
        rows = workbook[sheet_name].iter_rows(values_only=True)
        next(rows)
        ids.extend(str(row[0]) for row in rows)

    assert ids == ["term-1", "term-2", "term-3"]
    first_row = next(workbook["术语对照"].iter_rows(min_row=2, values_only=True))
    assert first_row[1] == "'=Formula"
    assert first_row[-1] == "bad\\u0000note"
    assert artifact.sheet_names == ("构建摘要", "术语对照", "同名异译", "人工调整记录", "术语对照-002")


def test_quality_renderer_uses_manifest_and_section_pages_without_loading_complete_snapshot(tmp_path: Path) -> None:
    repository = SqliteTerminologyRepository(
        TerminologyConnectionFactory(TerminologyPaths(tmp_path / "storage")),
        "project-1",
    )
    try:
        repository.put_build(build())
        snapshot = TerminologyReportSnapshotFactory(repository).freeze(build().ref, draft=draft())
        repository.put_report_snapshot(snapshot)
        statements: list[str] = []
        repository._connection.set_trace_callback(statements.append)

        artifact = QualityExcelRenderer(TerminologyReportQueryService(repository)).render(
            snapshot.ref,
            tmp_path / "tracked-quality.xlsx",
            page_size=1,
        )

        assert artifact.path.is_file()
        assert not any(
            "payload_json from report_snapshots" in statement.lower().replace("\n", " ") for statement in statements
        )
    finally:
        repository._connection.set_trace_callback(None)
        repository.close()


@pytest.mark.slow
def test_quality_renderer_streams_fifty_thousand_terms_from_sqlite_pages(tmp_path: Path) -> None:
    repository = SqliteTerminologyRepository(
        TerminologyConnectionFactory(TerminologyPaths(tmp_path / "storage")),
        "project-1",
    )
    try:
        repository.put_build(build())
        decisions = tuple(
            replace(
                decision(f"term-{index:05d}"),
                original=f"Term {index:05d}",
                normalized_original=f"term {index:05d}",
                translation=f"术语 {index:05d}",
            )
            for index in range(50_000)
        )
        snapshot = TerminologyReportSnapshotFactory(repository).freeze(
            build().ref,
            draft=TerminologyDraft(draft().ref, decisions),
        )
        repository.put_report_snapshot(snapshot)
        statements: list[str] = []
        repository._connection.set_trace_callback(statements.append)

        artifact = QualityExcelRenderer(TerminologyReportQueryService(repository)).render(
            snapshot.ref,
            tmp_path / "quality-50k.xlsx",
            page_size=1000,
        )

        workbook = load_workbook(artifact.path, read_only=True)
        assert sum(1 for _ in workbook["术语对照"].iter_rows()) == 50_001
        assert not any(
            "payload_json from report_snapshots" in statement.lower().replace("\n", " ") for statement in statements
        )
    finally:
        repository._connection.set_trace_callback(None)
        repository.close()
