from __future__ import annotations

from dataclasses import replace
from pathlib import Path

from openpyxl import load_workbook

from tests.application.terminology.story08_support import decision
from transbridge.application.terminology.diff import CanonicalDiffEngine
from transbridge.application.terminology.in_memory import InMemoryTerminologyRepository
from transbridge.application.terminology.models import TerminologyVersionRef
from transbridge.application.terminology.narrative import ChangeNarrativeProjector
from transbridge.application.terminology.renderers.changelog_excel import ChangeLogExcelRenderer
from transbridge.application.terminology.renderers.changelog_markdown import ChangeLogMarkdownRenderer


def test_markdown_and_excel_have_identical_semantic_manifest_from_one_frozen_document(tmp_path: Path) -> None:
    changed = decision()
    diff = CanonicalDiffEngine().compare(None, target_version_id="v1", decisions=(changed,))
    version_ref = TerminologyVersionRef("v1", "project-1", "variant-1", "version-content")
    document = ChangeNarrativeProjector().project(
        version_ref=version_ref,
        diff=diff,
        decisions=(changed,),
        conflicts=(),
        manual_actions=(),
        diagnostics=("published from frozen facts",),
    )
    repository = InMemoryTerminologyRepository()
    repository.put_changelog(document)

    markdown = ChangeLogMarkdownRenderer(repository).render(document.ref, tmp_path / "changes.md")
    excel = ChangeLogExcelRenderer(repository).render(document.ref, tmp_path / "changes.xlsx")

    assert markdown.semantic_manifest == excel.semantic_manifest
    assert markdown.semantic_manifest.change_count == len(document.changes)
    assert markdown.semantic_manifest.message_count == len(document.user_messages)
    assert document.changes[0].change_id in markdown.path.read_text(encoding="utf-8")
    workbook = load_workbook(excel.path, read_only=True)
    assert workbook.sheetnames == ["最终用户摘要", "维护者完整明细", "发布绑定事实"]
    assert sum(1 for _ in workbook["维护者完整明细"].iter_rows()) == len(document.changes) + 1


def test_changelog_rebuild_ignores_unrelated_current_state(tmp_path: Path) -> None:
    changed = decision(translation="=危险译名")
    diff = CanonicalDiffEngine().compare(None, target_version_id="v1", decisions=(changed,))
    version_ref = TerminologyVersionRef("v1", "project-1", "variant-1", "version-content")
    document = ChangeNarrativeProjector().project(
        version_ref=version_ref,
        diff=diff,
        decisions=(changed,),
        conflicts=(),
        manual_actions=(),
    )
    repository = InMemoryTerminologyRepository()
    repository.put_changelog(document)
    renderer = ChangeLogMarkdownRenderer(repository)

    first = renderer.render(document.ref, tmp_path / "first.md")
    repository.put_report_snapshot  # unrelated mutable repository capability must not be consulted
    second = renderer.render(document.ref, tmp_path / "second.md")

    assert first.semantic_manifest == second.semantic_manifest
    assert first.sha256 == second.sha256


class _TrackingPagedSource:
    def __init__(self, repository: InMemoryTerminologyRepository) -> None:
        self.repository = repository
        self.calls: list[str] = []

    def get_changelog(self, ref):
        raise AssertionError("renderer must not load the complete ChangeLogDocument")

    def __getattr__(self, name):
        if name == "get_changelog_manifest" or name.startswith("list_changelog_"):
            method = getattr(self.repository, name)

            def tracked(*args, **kwargs):
                self.calls.append(name)
                return method(*args, **kwargs)

            return tracked
        raise AttributeError(name)


def test_large_changelog_renderers_read_document_bound_pages(tmp_path: Path) -> None:
    decisions = (
        decision("term-1"),
        replace(decision("term-2"), original="Wyvern", normalized_original="wyvern", translation="飞龙"),
        replace(decision("term-3"), original="Whelp", normalized_original="whelp", translation="幼龙"),
    )
    diff = CanonicalDiffEngine().compare(None, target_version_id="v1", decisions=decisions)
    version_ref = TerminologyVersionRef("v1", "project-1", "variant-1", "version-content")
    document = ChangeNarrativeProjector().project(
        version_ref=version_ref,
        diff=diff,
        decisions=decisions,
        conflicts=(),
        manual_actions=(),
    )
    repository = InMemoryTerminologyRepository()
    repository.put_changelog(document)
    source = _TrackingPagedSource(repository)

    markdown = ChangeLogMarkdownRenderer(source).render(document.ref, tmp_path / "paged.md", page_size=1)
    excel = ChangeLogExcelRenderer(source).render(document.ref, tmp_path / "paged.xlsx", page_size=1)

    assert markdown.semantic_manifest == excel.semantic_manifest
    assert source.calls.count("list_changelog_changes") >= 6
    assert "get_changelog_manifest" in source.calls
