"""Contract tests for the canonical standalone polish report."""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import json

from openpyxl import load_workbook
import pytest

from transbridge.application.contracts import OperationOutcome
from transbridge.application.io import EntryKey, EntryRevision, SourceNamespace
from transbridge.application.translation import (
    build_polish_report_snapshot,
    postprocess_report as report_module,
    render_report_bundle,
)


@dataclass(frozen=True)
class _Entry:
    id: str
    original: str
    translation: str
    stage: int = 1
    context: str = "NPC_:FULL"
    revision: EntryRevision = EntryRevision(2)

    @property
    def identity(self) -> EntryKey:
        return EntryKey(SourceNamespace("fixture"), self.id)


def _snapshot():
    entries = (
        _Entry("accepted", "Hello", "你好"),
        _Entry("rejected", "Goodbye", "再见"),
        _Entry("failed", "Wait", "等等"),
    )
    results = {
        "accepted": {
            "original_translation": "你好",
            "polished_translation": "您好",
            "confidence": 0.92,
            "needs_arbitration": False,
            "note": "更自然",
            "verdict": "pass",
            "refined_translation": "你好呀",
            "changes": [{"aspect": "fluency", "before": "你好", "after": "您好"}],
            "issues": (),
        },
        "rejected": {
            "original_translation": "再见",
            "polished_translation": "告辞",
            "confidence": 0.7,
            "needs_arbitration": True,
            "note": "语气变化过大",
            "verdict": "reject",
            "changes": [{"aspect": "tone", "reason": "更古雅"}],
            "issues": (
                {
                    "issue_type": "tone_risk",
                    "severity": "warning",
                    "message": "Tone changed.",
                    "suggestion": "Review manually.",
                },
            ),
        },
        "failed": {
            "original_translation": "等等",
            "polished_translation": "等等",
            "confidence": float("nan"),
            "needs_arbitration": True,
            "note": "LLM unavailable",
            "verdict": "pending",
            "changes": (),
        },
    }
    return build_polish_report_snapshot(
        results,
        entries,
        accepted_entry_ids={"accepted"},
        rejected_entry_ids={"rejected"},
        failed_entry_ids={"failed"},
        run_id="polish-run",
        polish_level="moderate",
        run_spec_summary={"stages": ["check", "refine", "polish", "arbitrate"]},
    )


def test_polish_snapshot_preserves_decision_and_report_details() -> None:
    snapshot = _snapshot()

    assert snapshot.schema == "transbridge.polish-report.v1"
    assert snapshot.outcome is OperationOutcome.PARTIAL
    assert snapshot.input_count == 3
    assert snapshot.accepted_count == 1
    assert snapshot.failure_count == 1
    assert snapshot.run_spec_summary["polish_counts"] == {"accepted": 1, "rejected": 1, "failed": 1}
    assert snapshot.run_spec_summary["polish_level"] == "moderate"

    entries = {candidate.entry_key.local_key: candidate.to_dict() for candidate in snapshot.candidates}
    accepted = entries["accepted"]["report_details"]
    assert accepted["result_status"] == "accepted"
    assert accepted["confidence"] == 0.92
    assert accepted["refined_translation"] == "你好呀"
    assert accepted["changes"][0]["aspect"] == "fluency"
    assert entries["accepted"]["phases"] == ["check", "refine", "polish", "arbitrate"]
    assert entries["rejected"]["report_details"]["verdict"] == "reject"
    assert entries["failed"]["report_details"]["confidence"] == 0.0
    assert any(diagnostic.code == "POLISH_ISSUE_TONE_RISK" for diagnostic in snapshot.diagnostics)
    assert any(diagnostic.code == "POLISH_ENTRY_FAILED" for diagnostic in snapshot.diagnostics)
    json.dumps(snapshot.to_dict(), ensure_ascii=False, allow_nan=False)


def test_explicit_pending_candidate_is_neither_rejected_nor_failed() -> None:
    entry = _Entry("pending", "Hello", "你好")
    snapshot = build_polish_report_snapshot(
        {"pending": {"polished_translation": "您好", "confidence": 1, "accepted": True}},
        [entry],
        run_id="pending-run",
        pending_entry_ids={"pending"},
    )
    assert snapshot.accepted_count == snapshot.failure_count == 0
    assert dict(snapshot.candidates[0].report_details)["result_status"] == "pending"
    assert snapshot.run_spec_summary["polish_counts"]["pending"] == 1
    assert snapshot.diagnostics == ()
    with pytest.raises(ValueError, match="disjoint"):
        build_polish_report_snapshot(
            {}, [entry], run_id="invalid", pending_entry_ids={"pending"}, accepted_entry_ids={"pending"}
        )


def test_polish_bundle_keeps_details_in_all_three_formats(tmp_path) -> None:
    result = render_report_bundle(_snapshot(), base_dir=tmp_path)

    assert result.outcome is OperationOutcome.COMPLETED
    assert result.value is not None
    assert {artifact.renderer for artifact in result.value.artifacts} == {"json", "csv", "excel"}
    assert all(artifact.artifact_path for artifact in result.value.artifacts)

    artifacts = {artifact.renderer: artifact for artifact in result.value.artifacts}
    payload = json.loads(artifacts["json"].content)
    assert payload["entries"][0]["report_details"]["note"] == "更自然"
    csv_text = artifacts["csv"].content.decode("utf-8")
    assert "result_status" in csv_text and "更自然" in csv_text and '""aspect"":""fluency""' in csv_text

    workbook = load_workbook(BytesIO(artifacts["excel"].content), read_only=True)
    rows = tuple(workbook["Entries"].iter_rows(values_only=True))
    headers = rows[0]
    accepted_row = dict(zip(headers, rows[1], strict=True))
    assert accepted_row["result_status"] == "accepted"
    assert accepted_row["confidence"] == 0.92
    assert "fluency" in accepted_row["changes"]


def test_polish_bundle_rotates_each_format_independently(tmp_path) -> None:
    for suffix in ("json", "csv", "xlsx"):
        for index in range(22):
            path = tmp_path / f"polish-report-seed-{index:02d}.{suffix}"
            path.write_text("seed", encoding="utf-8")
            path.touch()

    result = render_report_bundle(_snapshot(), base_dir=tmp_path, max_files_per_format=20)

    assert result.outcome is OperationOutcome.COMPLETED
    for suffix in ("json", "csv", "xlsx"):
        assert len(tuple(tmp_path.glob(f"polish-report-*.{suffix}"))) == 20


def test_polish_bundle_reports_rotation_failure_without_losing_artifacts(tmp_path, monkeypatch) -> None:
    def fail_rotation(*_args, **_kwargs):
        raise PermissionError("locked history")

    monkeypatch.setattr(report_module, "_rotate_artifacts", fail_rotation)
    result = render_report_bundle(_snapshot(), base_dir=tmp_path)

    assert result.outcome is OperationOutcome.COMPLETED
    assert result.value is not None
    assert result.value.produced == 3
    assert result.value.failed == 0
    assert result.counts.failed == 0
    assert len([diagnostic for diagnostic in result.diagnostics if diagnostic.code == "REPORT_ROTATION_FAILED"]) == 3
    assert all(diagnostic.severity.value == "warning" for diagnostic in result.diagnostics)


def test_polish_status_collections_must_be_disjoint() -> None:
    entry = _Entry("same", "Source", "Draft")
    try:
        build_polish_report_snapshot(
            {},
            (entry,),
            accepted_entry_ids={"same"},
            rejected_entry_ids={"same"},
            run_id="bad-status",
        )
    except ValueError as exc:
        assert "disjoint" in str(exc)
    else:  # pragma: no cover - assertion guard
        raise AssertionError("overlapping status collections must fail")


def test_rejecting_every_candidate_is_a_completed_review_not_a_run_failure() -> None:
    entry = _Entry("rejected", "Source", "Draft")
    snapshot = build_polish_report_snapshot(
        {
            "rejected": {
                "original_translation": "Draft",
                "polished_translation": "Candidate",
                "confidence": 0.8,
                "needs_arbitration": True,
                "changes": (),
            }
        },
        (entry,),
        rejected_entry_ids={"rejected"},
        run_id="all-rejected",
    )

    assert snapshot.outcome is OperationOutcome.COMPLETED
    assert snapshot.accepted_count == 0
    assert snapshot.failure_count == 0


def test_polish_report_details_redact_secrets() -> None:
    entry = _Entry("secret", "Source", "Draft")
    snapshot = build_polish_report_snapshot(
        {
            "secret": {
                "original_translation": "Draft",
                "polished_translation": "Candidate",
                "confidence": 0.9,
                "needs_arbitration": False,
                "note": "upstream token=super-secret-token-value",
                "changes": (),
            }
        },
        (entry,),
        run_id="secret-redaction",
    )

    rendered = json.dumps(snapshot.to_dict(), ensure_ascii=False)
    assert "super-secret-token-value" not in rendered
    assert "***REDACTED***" in rendered
