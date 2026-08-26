"""Canonical report contracts: golden parity, safe render failure and secret-free output."""

from __future__ import annotations

from dataclasses import replace
from io import BytesIO
import json

from openpyxl import load_workbook

from transbridge.application.contracts import Diagnostic, OperationOutcome
from transbridge.application.io import EntryKey, EntryRevision, SourceNamespace, StagePolicy
from transbridge.application.translation import (
    CsvReportRenderer,
    ExcelReportRenderer,
    InMemoryPostProcessCheckpointPort,
    JsonReportRenderer,
    PostProcessStageOutcome,
    PostProcessWorkload,
    ReportSnapshot,
    TranslationInput,
    render_report,
)


def _refine(candidates):
    return PostProcessStageOutcome(
        "refine", tuple(candidate.with_text("refined-text", "refine") for candidate in candidates)
    )


def _run_snapshot(*, run_id: str = "golden-run") -> ReportSnapshot:
    workload = PostProcessWorkload(
        (_refine,),
        stage_policy=StagePolicy(),
        stage_names=("refine",),
        checkpoint_port=InMemoryPostProcessCheckpointPort(),
    )
    entry = TranslationInput(
        EntryKey(SourceNamespace("fixture"), "golden"), EntryRevision(3), "source-text", "draft-text", 2
    )
    result = workload.run(
        run_id,
        (entry,),
        owner_id="owner",
        run_spec_summary={"target_locale": "zh-CN", "config_revision": 7},
    )
    assert result.outcome is OperationOutcome.COMPLETED
    assert result.value is not None
    return result.value


def test_golden_counts_are_consistent_across_all_renderers(tmp_path) -> None:
    snapshot = _run_snapshot()
    result = render_report(
        snapshot,
        (JsonReportRenderer(), CsvReportRenderer(), ExcelReportRenderer()),
        base_dir=tmp_path,
    )

    assert result.outcome is OperationOutcome.COMPLETED
    assert result.value is not None
    assert result.value.produced == 3
    assert result.value.failed == 0

    parsed = json.loads(result.value.artifacts[0].content)
    assert parsed["schema"] == "transbridge.postprocess-report.v1"
    assert parsed["run_id"] == "golden-run"
    assert parsed["counts"]["input"] == 1
    assert parsed["counts"]["accepted"] == 1
    assert parsed["issues"] == 0
    assert parsed["failures"] == 0
    assert parsed["timing_ms"][0][0] == "refine"
    assert parsed["run_spec_summary"]["target_locale"] == "zh-CN"

    csv_text = result.value.artifacts[1].content.decode("utf-8")
    assert "golden-run" in csv_text and "refined-text" in csv_text
    assert csv_text.count("\n") == 2  # header + one entry

    excel_bytes = result.value.artifacts[2].content
    assert excel_bytes[:2] == b"PK"  # xlsx is a zip container
    workbook = load_workbook(BytesIO(excel_bytes), read_only=True)
    assert workbook.sheetnames == ["Summary", "Entries", "Diagnostics", "Stages"]
    assert tuple(workbook["Entries"].iter_rows(min_row=2, max_row=2, values_only=True))[0][2:5] == (
        "source-text",
        "draft-text",
        "refined-text",
    )
    assert dict(workbook["Summary"].iter_rows(values_only=True))["run_spec.target_locale"] == "zh-CN"
    artifacts = {artifact.renderer: artifact for artifact in result.value.artifacts}
    assert all(artifact.sha256 for artifact in artifacts.values())
    assert all(artifact.artifact_path for artifact in artifacts.values())


def test_renderer_failure_never_rolls_back_and_carries_report_diagnostic(tmp_path) -> None:
    snapshot = _run_snapshot(run_id="render-fail-run")

    class BrokenRenderer:
        name = "broken"

        def render(self, snapshot, *, base_dir=None):
            raise OSError("disk full")

    result = render_report(snapshot, (JsonReportRenderer(), BrokenRenderer()), base_dir=tmp_path)

    assert result.outcome is OperationOutcome.PARTIAL
    assert result.value is not None
    assert result.value.produced == 1
    assert result.value.failed == 1
    assert any(d.code == "REPORT_RENDER_FAILED" for d in result.diagnostics)
    assert result.counts.failed == 1


def test_report_never_contains_prompts_or_secrets() -> None:
    snapshot = _run_snapshot(run_id="secret-scan")
    rendered = json.dumps(snapshot.to_dict(), ensure_ascii=False)
    for forbidden in ("Prompt", "bearer", "sk-", "api_key", "Authorization"):
        assert forbidden not in rendered
    json_artifact = JsonReportRenderer().render(snapshot).content.decode("utf-8")
    assert "Authorization" not in json_artifact


def test_snapshot_boundary_recursively_detaches_and_redacts_metadata_and_diagnostics() -> None:
    summary = {
        "model": "fixture-model",
        "api_key": "sk-abcdefghijklmnopqrstuvwxyz123456",
        "nested": {
            "prompt": "system instructions",
            "ordinary": "keep-me",
            "items": [{"token": "sensitive-token-value", "label": "keep-label"}],
        },
    }
    diagnostic_details = {
        "authorization": "Bearer abcdefghijklmnopqrstuvwxyz123456",
        "ordinary": {"secret": "hidden-value", "label": "keep-detail"},
    }
    snapshot = replace(
        _run_snapshot(run_id="secret-boundary"),
        run_spec_summary=summary,
        diagnostics=(
            Diagnostic(
                "UPSTREAM_FAILED",
                "upstream token=super-secret-token-value",
                details=tuple(diagnostic_details.items()),
            ),
        ),
    )
    summary["nested"]["ordinary"] = "mutated-after-construction"
    diagnostic_details["ordinary"]["label"] = "mutated-after-construction"

    payload = snapshot.to_dict()
    assert payload["run_spec_summary"]["model"] == "fixture-model"
    assert payload["run_spec_summary"]["api_key"] == "***REDACTED***"
    assert payload["run_spec_summary"]["nested"]["prompt"] == "***REDACTED***"
    assert payload["run_spec_summary"]["nested"]["ordinary"] == "keep-me"
    assert payload["run_spec_summary"]["nested"]["items"][0] == {
        "token": "***REDACTED***",
        "label": "keep-label",
    }
    diagnostic = payload["diagnostics"][0]
    assert "super-secret-token-value" not in diagnostic["message"]
    assert diagnostic["details"]["authorization"] == "***REDACTED***"
    assert diagnostic["details"]["ordinary"] == {
        "secret": "***REDACTED***",
        "label": "keep-detail",
    }
    json.dumps(payload, ensure_ascii=False, allow_nan=False)


def test_report_snapshot_fingerprint_is_stable_per_run_id() -> None:
    first = _run_snapshot(run_id="fp-run")
    second = _run_snapshot(run_id="fp-run")
    assert first.fingerprint == second.fingerprint
    third = _run_snapshot(run_id="fp-run-other")
    assert first.fingerprint != third.fingerprint


def test_spreadsheet_renderers_escape_untrusted_formula_text() -> None:
    snapshot = _run_snapshot(run_id="formula-safe")
    candidate = replace(
        snapshot.candidates[0],
        original='=HYPERLINK("https://invalid")',
        before_text="+cmd",
        text="@SUM(A1:A2)",
        context="-1+1",
        report_details=(("note", '=WEBSERVICE("https://invalid")'),),
    )
    snapshot = replace(
        snapshot,
        candidates=(candidate,),
        diagnostics=(Diagnostic("FORMULA_INPUT", "+SUM(A1:A2)"),),
    )

    csv_text = CsvReportRenderer().render(snapshot).content.decode("utf-8")
    assert "'+cmd" in csv_text
    assert "'@SUM" in csv_text
    assert "'=WEBSERVICE" in csv_text

    workbook = load_workbook(BytesIO(ExcelReportRenderer().render(snapshot).content), data_only=False)
    row = tuple(workbook["Entries"].iter_rows(min_row=2, max_row=2, values_only=False))[0]
    assert row[2].value.startswith("'=") and row[2].data_type != "f"
    assert row[3].value.startswith("'+") and row[3].data_type != "f"
    assert row[4].value.startswith("'@") and row[4].data_type != "f"
    assert row[8].value.startswith("'-") and row[8].data_type != "f"
    entry_headers = [cell.value for cell in workbook["Entries"][1]]
    note_cell = row[entry_headers.index("note")]
    assert note_cell.value.startswith("'=") and note_cell.data_type != "f"
    diagnostic_row = tuple(workbook["Diagnostics"].iter_rows(min_row=2, max_row=2, values_only=False))[0]
    diagnostic_headers = [cell.value for cell in workbook["Diagnostics"][1]]
    message_cell = diagnostic_row[diagnostic_headers.index("message")]
    assert message_cell.value.startswith("'+") and message_cell.data_type != "f"
