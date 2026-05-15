"""文本类文件解析器：Excel / CSV / Markdown / TXT / JSON。"""

import csv
import io
import json
import re
from pathlib import Path

from .base import FileParser, ParsedDocument


class TextFileParser(FileParser):
    supported_extensions = [".xlsx", ".csv", ".md", ".txt", ".json"]

    def parse(self, path: Path) -> ParsedDocument:
        ext = path.suffix.lower()
        if ext == ".xlsx":
            return self._parse_xlsx(path)
        elif ext == ".csv":
            return self._parse_csv(path)
        elif ext == ".md":
            return self._parse_markdown(path)
        elif ext == ".json":
            return self._parse_json(path)
        else:
            return self._parse_text(path)

    # ── 各格式解析 ───────────────────────────────────────

    def _parse_xlsx(self, path: Path) -> ParsedDocument:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sections = []
            raw_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None for c in row):
                        row_vals = [str(c) if c is not None else "" for c in row]
                        rows.append(row_vals)
                        raw_parts.append("\t".join(row_vals))
                if rows:
                    sections.append({"heading": sheet_name, "rows": rows})
            sheet_names = list(wb.sheetnames)
            return ParsedDocument(
                source_path=path, format="excel", title=path.stem,
                sections=sections, raw_text="\n".join(raw_parts),
                metadata={"sheets": sheet_names},
            )
        finally:
            wb.close()

    def _parse_csv(self, path: Path) -> ParsedDocument:
        rows = []
        raw_parts = []
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            if headers:
                rows.append(list(headers))
            for row in reader:
                rows.append(list(row))
                raw_parts.append("\t".join(row))
        return ParsedDocument(
            source_path=path, format="csv", title=path.stem,
            sections=[{"heading": path.stem, "rows": rows}],
            raw_text="\n".join(raw_parts),
        )

    _MD_HEADING_RE = re.compile(r"^(#{1,6})\s+(.*)")

    def _parse_markdown(self, path: Path) -> ParsedDocument:
        text = path.read_text(encoding="utf-8", errors="replace")
        sections = []
        current_heading = ""
        current_content: list[str] = []
        for line in text.split("\n"):
            m = self._MD_HEADING_RE.match(line)
            if m:
                if current_heading or current_content:
                    sections.append({"heading": current_heading, "content": "\n".join(current_content)})
                current_heading = m.group(2).strip()
                current_content = []
            else:
                current_content.append(line)
        if current_heading or current_content:
            sections.append({"heading": current_heading, "content": "\n".join(current_content)})
        return ParsedDocument(
            source_path=path, format="markdown", title=path.stem,
            sections=sections, raw_text=text,
        )

    def _parse_text(self, path: Path) -> ParsedDocument:
        text = path.read_text(encoding="utf-8", errors="replace")
        return ParsedDocument(
            source_path=path, format="text", title=path.stem,
            sections=[{"heading": path.stem, "content": text}],
            raw_text=text,
        )

    def _parse_json(self, path: Path) -> ParsedDocument:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        raw_text = json.dumps(data, ensure_ascii=False, indent=2)
        return ParsedDocument(
            source_path=path, format="json", title=path.stem,
            raw_text=raw_text,
            metadata=data if isinstance(data, dict) else {"data": data},
        )
