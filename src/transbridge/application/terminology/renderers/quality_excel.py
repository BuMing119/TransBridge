"""Write-only Excel projection of one frozen terminology quality snapshot."""

from __future__ import annotations

from dataclasses import dataclass
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet._write_only import WriteOnlyWorksheet

from ..identity import canonical_digest
from ..models import ConflictGroup, ManualAction, TermDecision, TerminologyReportSnapshotRef
from ..report_queries import ReportSnapshotSummary, TerminologyReportQueryService
from ._artifact import ArtifactPublishPolicy, publish_staged
from ._manifest import RenderedArtifact, SemanticManifest
from .spreadsheet_safety import expanded_spreadsheet_rows

RENDERER_VERSION = "terminology-quality-excel.v1"
EXCEL_MAX_ROWS = 1_048_576

SUMMARY_SHEET = "构建摘要"
TERMS_SHEET = "术语对照"
CONFLICTS_SHEET = "同名异译"
MANUAL_SHEET = "人工调整记录"

TERMS_HEADERS = (
    "术语ID",
    "原名",
    "译名",
    "规范原名",
    "作用域",
    "状态",
    "已抑制",
    "证据数",
    "备注",
)
CONFLICT_HEADERS = (
    "冲突组ID",
    "规范原名",
    "风险",
    "状态",
    "推荐译名",
    "译名变体",
    "备注",
)
MANUAL_HEADERS = (
    "操作ID",
    "术语ID",
    "操作类型",
    "操作者",
    "发生时间",
    "基线版本",
    "操作前摘要",
    "操作后摘要",
    "原因",
    "替换术语ID",
)


@dataclass(slots=True)
class _SheetSegment:
    name: str
    data_rows: int = 0
    first_stable_id: str | None = None
    last_stable_id: str | None = None


class _LogicalSheetWriter:
    def __init__(self, workbook: Workbook, title: str, headers: tuple[str, ...], *, max_rows: int) -> None:
        if max_rows < 2:
            raise ValueError("Excel sheet row capacity must leave room for a header and one data row")
        self._workbook = workbook
        self._title = title
        self._headers = headers
        self._max_rows = min(max_rows, EXCEL_MAX_ROWS)
        self._sheet: WriteOnlyWorksheet
        self._rows = 0
        self.segments: list[_SheetSegment] = []
        self._open_segment()

    def append(self, values: tuple[object, ...], *, stable_id: str) -> None:
        for physical_row in expanded_spreadsheet_rows(values):
            if self._rows >= self._max_rows:
                self._open_segment()
            self._sheet.append(physical_row)
            self._rows += 1
            segment = self.segments[-1]
            segment.data_rows += 1
            segment.first_stable_id = segment.first_stable_id or stable_id
            segment.last_stable_id = stable_id

    def _open_segment(self) -> None:
        index = len(self.segments) + 1
        suffix = "" if index == 1 else f"-{index:03d}"
        title = f"{self._title[: 31 - len(suffix)]}{suffix}"
        self._sheet = self._workbook.create_sheet(title)
        self._sheet.append(self._headers)
        self._rows = 1
        self.segments.append(_SheetSegment(title))


class QualityExcelRenderer:
    format = "xlsx"
    renderer_version = RENDERER_VERSION

    def __init__(self, queries: TerminologyReportQueryService, *, max_sheet_rows: int = EXCEL_MAX_ROWS) -> None:
        if isinstance(max_sheet_rows, bool) or not isinstance(max_sheet_rows, int):
            raise ValueError("Excel sheet row capacity must be an integer")
        self._queries = queries
        self._max_sheet_rows = max_sheet_rows

    def render(
        self,
        ref: TerminologyReportSnapshotRef,
        target: str | Path,
        *,
        policy: ArtifactPublishPolicy = ArtifactPublishPolicy.FAIL_IF_EXISTS,
        page_size: int = 1000,
    ) -> RenderedArtifact:
        summary = self._queries.summary(ref)
        workbook = Workbook(write_only=True)
        summary_sheet = workbook.create_sheet(SUMMARY_SHEET)
        terms = _LogicalSheetWriter(workbook, TERMS_SHEET, TERMS_HEADERS, max_rows=self._max_sheet_rows)
        conflicts = _LogicalSheetWriter(workbook, CONFLICTS_SHEET, CONFLICT_HEADERS, max_rows=self._max_sheet_rows)
        manual = _LogicalSheetWriter(workbook, MANUAL_SHEET, MANUAL_HEADERS, max_rows=self._max_sheet_rows)

        for item in self._queries.iter_terms(ref, page_size=page_size):
            terms.append(_term_row(item), stable_id=item.term_id)
        for item in self._queries.iter_conflicts(ref, page_size=page_size):
            conflicts.append(_conflict_row(item), stable_id=item.conflict_group_id)
        for item in self._queries.iter_manual_actions(ref, page_size=page_size):
            manual.append(_manual_row(item), stable_id=item.action_id)

        all_segments = (
            (TERMS_SHEET, terms.segments),
            (CONFLICTS_SHEET, conflicts.segments),
            (MANUAL_SHEET, manual.segments),
        )
        _write_summary(summary_sheet, summary, all_segments)
        sheet_names = tuple(sheet.title for sheet in workbook.worksheets)
        published = publish_staged(target, workbook.save, policy=policy)
        manifest = SemanticManifest(
            ref.snapshot_id,
            ref.content_digest,
            ref.content_digest,
            canonical_digest(summary, namespace="terminology.quality-summary.v1"),
            canonical_digest(
                tuple(
                    (logical, tuple((item.name, item.data_rows) for item in segments))
                    for logical, segments in all_segments
                ),
                namespace="terminology.quality-sheet-manifest.v1",
            ),
            0,
            summary.term_count + summary.conflict_count + summary.manual_action_count,
            0,
        )
        return RenderedArtifact(
            self.format,
            self.renderer_version,
            published.path,
            published.size,
            published.sha256,
            manifest,
            sheet_names,
        )


def _write_summary(
    sheet: WriteOnlyWorksheet,
    summary: ReportSnapshotSummary,
    sections: tuple[tuple[str, list[_SheetSegment]], ...],
) -> None:
    rows: tuple[tuple[object, ...], ...] = (
        ("字段", "值"),
        ("snapshot_id", summary.snapshot_ref.snapshot_id),
        ("snapshot_digest", summary.snapshot_ref.content_digest),
        ("build_key", summary.build_ref.build_key),
        ("build_digest", summary.build_ref.content_digest),
        ("project_id", summary.project_id),
        ("variant_id", summary.variant_id),
        ("draft_identity", summary.draft_identity),
        ("source_count", summary.build_summary.source_count),
        ("evidence_count", summary.build_summary.evidence_count),
        ("candidate_count", summary.build_summary.candidate_count),
        ("conflict_count", summary.build_summary.conflict_count),
        ("excluded_count", summary.build_summary.excluded_count),
        ("report_term_count", summary.term_count),
        ("report_conflict_count", summary.conflict_count),
        ("manual_action_count", summary.manual_action_count),
    )
    for row in rows:
        for physical in expanded_spreadsheet_rows(row):
            sheet.append(physical)
    sheet.append(())
    sheet.append(("逻辑表", "工作表", "物理数据行数", "首个稳定ID", "末个稳定ID"))
    for logical, segments in sections:
        for segment in segments:
            sheet.append(
                tuple(
                    item
                    for item in expanded_spreadsheet_rows((
                        logical,
                        segment.name,
                        segment.data_rows,
                        segment.first_stable_id or "",
                        segment.last_stable_id or "",
                    ))[0]
                )
            )


def _term_row(item: TermDecision) -> tuple[object, ...]:
    return (
        item.term_id,
        item.original,
        item.translation,
        item.normalized_original,
        item.scope.canonical_key,
        item.status.value,
        item.suppressed,
        len(item.evidence_ids),
        item.notes,
    )


def _conflict_row(item: ConflictGroup) -> tuple[object, ...]:
    variants = tuple(
        {
            "translation": variant.normalized_translation,
            "candidate_ids": variant.candidate_ids,
            "evidence_ids": variant.evidence_ids,
        }
        for variant in item.variants
    )
    return (
        item.conflict_group_id,
        item.normalized_original,
        item.risk.value,
        item.status.value,
        item.recommended_translation or "",
        json.dumps(variants, ensure_ascii=False, sort_keys=True, separators=(",", ":")),
        item.notes,
    )


def _manual_row(item: ManualAction) -> tuple[object, ...]:
    return (
        item.action_id,
        item.term_id,
        item.action_type.value,
        item.actor,
        item.occurred_at,
        item.base_version_id or "",
        item.before_digest or "",
        item.after_digest or "",
        item.reason or "",
        item.replacement_term_id or "",
    )


__all__ = [
    "CONFLICT_HEADERS",
    "CONFLICTS_SHEET",
    "EXCEL_MAX_ROWS",
    "MANUAL_HEADERS",
    "MANUAL_SHEET",
    "QualityExcelRenderer",
    "RENDERER_VERSION",
    "SUMMARY_SHEET",
    "TERMS_HEADERS",
    "TERMS_SHEET",
]
