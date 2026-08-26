"""Canonical report renderers for ReportSnapshot.

UI, Excel and durable history all consume the same ReportSnapshot; a renderer
is a pure projection. Renderer failures never roll back already-committed
business: :func:`render_report` returns a typed partial/failed result carrying
a REPORT_RENDER_FAILED diagnostic instead of raising.
Renderers receive only the snapshot — prompts, credentials and model messages
never enter report content.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
import hashlib
import io
import json
from pathlib import Path
from typing import Protocol

from transbridge.application.contracts import (
    Diagnostic,
    DiagnosticSeverity,
    ErrorCategory,
    OperationCounts,
    OperationOutcome,
    OperationResult,
)

from .postprocess import ReportSnapshot


@dataclass(frozen=True, slots=True)
class ReportRenderResult:
    renderer: str
    size: int
    sha256: str
    artifact_path: str | None = None
    content: bytes = b""

    @property
    def fingerprint(self) -> str:
        return self.sha256


class ReportRendererPort(Protocol):
    name: str

    def render(self, snapshot: ReportSnapshot, *, base_dir: Path | None = None) -> ReportRenderResult: ...


@dataclass(frozen=True, slots=True)
class ReportRenderOutcome:
    artifacts: tuple[ReportRenderResult, ...]
    produced: int
    failed: int
    run_id: str


class JsonReportRenderer:
    name = "json"

    def render(self, snapshot: ReportSnapshot, *, base_dir: Path | None = None) -> ReportRenderResult:
        content = json.dumps(
            snapshot.to_dict(),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
        return ReportRenderResult(
            self.name,
            len(content),
            _sha256(content),
            artifact_path=_artifact(base_dir, _report_stem(snapshot), "json", content),
            content=content,
        )


class CsvReportRenderer:
    name = "csv"

    def render(self, snapshot: ReportSnapshot, *, base_dir: Path | None = None) -> ReportRenderResult:
        stream = io.StringIO()
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow((
            "schema",
            "run_id",
            "outcome",
            "namespace",
            "local_key",
            "stage",
            "accepted",
            "before",
            "candidate",
            "result_status",
            "confidence",
            "needs_arbitration",
            "verdict",
            "refined_translation",
            "changes",
            "note",
            "issues",
        ))
        for candidate in snapshot.candidates:
            key = candidate.entry_key.to_dict()
            details = dict(candidate.report_details)
            writer.writerow(
                tuple(
                    _spreadsheet_value(value)
                    for value in (
                        snapshot.schema,
                        snapshot.run_id,
                        snapshot.outcome.value,
                        key.get("namespace", ""),
                        key.get("local_key", ""),
                        candidate.stage,
                        "1" if candidate.accepted else "0",
                        candidate.before_text,
                        candidate.text,
                        details.get("result_status", ""),
                        details.get("confidence", ""),
                        "1" if details.get("needs_arbitration") else "0",
                        details.get("verdict", ""),
                        details.get("refined_translation", ""),
                        _csv_json(details.get("changes", ())),
                        details.get("note", ""),
                        _csv_json(details.get("issues", ())),
                    )
                )
            )
        content = stream.getvalue().encode("utf-8")
        return ReportRenderResult(
            self.name,
            len(content),
            _sha256(content),
            artifact_path=_artifact(base_dir, _report_stem(snapshot), "csv", content),
            content=content,
        )


class ExcelReportRenderer:
    """Additive renderer; keeps the canonical snapshot as the only source."""

    name = "excel"

    def render(self, snapshot: ReportSnapshot, *, base_dir: Path | None = None) -> ReportRenderResult:
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
        except ImportError as exc:  # pragma: no cover - degraded environment probe
            raise RuntimeError("openpyxl is unavailable; Excel rendering is degraded") from exc

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(("schema", _excel_value(snapshot.schema)))
        summary.append(("run_id", _excel_value(snapshot.run_id)))
        summary.append(("outcome", _excel_value(snapshot.outcome.value)))
        summary.append(("input_count", snapshot.input_count))
        summary.append(("accepted_count", snapshot.accepted_count))
        summary.append(("issue_count", snapshot.issue_count))
        summary.append(("failure_count", snapshot.failure_count))
        for key, value in sorted(snapshot.run_spec_summary.items()):
            summary.append((f"run_spec.{key}", _excel_value(value)))

        entries = workbook.create_sheet("Entries")
        entries.append((
            "namespace",
            "local_key",
            "original",
            "before",
            "candidate",
            "stage",
            "accepted",
            "phases",
            "context",
            "result_status",
            "confidence",
            "needs_arbitration",
            "verdict",
            "refined_translation",
            "changes",
            "note",
            "issues",
        ))
        for candidate in snapshot.candidates:
            key = candidate.entry_key.to_dict()
            details = dict(candidate.report_details)
            entries.append(
                tuple(
                    _excel_value(value)
                    for value in (
                        key.get("namespace", ""),
                        key.get("local_key", ""),
                        candidate.original,
                        candidate.before_text,
                        candidate.text,
                        candidate.stage,
                        candidate.accepted,
                        ", ".join(candidate.phases),
                        candidate.context,
                        details.get("result_status", ""),
                        details.get("confidence", ""),
                        details.get("needs_arbitration", ""),
                        details.get("verdict", ""),
                        details.get("refined_translation", ""),
                        details.get("changes", ()),
                        details.get("note", ""),
                        details.get("issues", ()),
                    )
                )
            )

        diagnostics = workbook.create_sheet("Diagnostics")
        diagnostics.append(("entry_id", "code", "severity", "category", "message", "retryable"))
        for diagnostic in snapshot.diagnostics:
            details = dict(diagnostic.details)
            entry_key = details.get("entry_key")
            entry_id = entry_key.get("local_key", "") if isinstance(entry_key, dict) else ""
            diagnostics.append(
                tuple(
                    _excel_value(value)
                    for value in (
                        entry_id,
                        diagnostic.code,
                        diagnostic.severity.value,
                        diagnostic.category.value if diagnostic.category else "",
                        diagnostic.message,
                        diagnostic.retryable,
                    )
                )
            )

        stages = workbook.create_sheet("Stages")
        stages.append(("phase", "duration_ms", "entry_count", "diagnostic_count"))
        for outcome in snapshot.stage_outcomes:
            stages.append(
                tuple(
                    _excel_value(value)
                    for value in (
                        outcome.phase,
                        outcome.duration_ms,
                        len(outcome.candidates),
                        len(outcome.diagnostics),
                    )
                )
            )

        stream = io.BytesIO()
        workbook.save(stream)
        content = stream.getvalue()
        return ReportRenderResult(
            self.name,
            len(content),
            _sha256(content),
            artifact_path=_artifact(base_dir, _report_stem(snapshot), "xlsx", content),
            content=content,
        )


def _artifact(base_dir: Path | None, stem: str, suffix: str, content: bytes) -> str | None:
    if base_dir is None:
        return None
    base_dir.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(content).hexdigest()[:16]
    target = base_dir / f"{stem}-{digest}.{suffix}"
    target.write_bytes(content)
    return str(target)


def default_report_renderers() -> tuple[ReportRendererPort, ...]:
    return (JsonReportRenderer(), CsvReportRenderer())


def render_report_bundle(
    snapshot: ReportSnapshot,
    *,
    base_dir: Path | None = None,
    max_files_per_format: int = 20,
) -> OperationResult[ReportRenderOutcome]:
    """Render the canonical JSON/CSV/Excel bundle and rotate durable history."""
    if max_files_per_format < 1:
        raise ValueError("max_files_per_format must be at least one")
    result = render_report(
        snapshot,
        (JsonReportRenderer(), CsvReportRenderer(), ExcelReportRenderer()),
        base_dir=base_dir,
    )
    rotation_diagnostics: list[Diagnostic] = []
    if base_dir is not None and result.value is not None:
        stem = _report_stem(snapshot)
        for suffix in ("json", "csv", "xlsx"):
            try:
                _rotate_artifacts(base_dir, stem, suffix, keep=max_files_per_format)
            except OSError as exc:
                rotation_diagnostics.append(
                    Diagnostic(
                        "REPORT_ROTATION_FAILED",
                        "A report history rotation failed; newly rendered artifacts are unaffected.",
                        category=ErrorCategory.INTERNAL,
                        severity=DiagnosticSeverity.WARNING,
                        details=(("format", suffix), ("error_type", type(exc).__name__)),
                    )
                )
    if rotation_diagnostics and result.value is not None:
        return OperationResult(
            result.outcome,
            result.value,
            diagnostics=(*result.diagnostics, *rotation_diagnostics),
            counts=result.counts,
            artifact_refs=result.artifact_refs,
            run_id=result.run_id,
        )
    return result


def render_report(
    snapshot: ReportSnapshot,
    renderers: tuple[ReportRendererPort, ...] | None = None,
    *,
    base_dir: Path | None = None,
) -> OperationResult[ReportRenderOutcome]:
    """Render all projections; never roll back committed business.

    A renderer exception is recorded as a REPORT_RENDER_FAILED diagnostic and
    a failed count. The render operation becomes PARTIAL or FAILED, while the
    already-committed translation operation remains unaffected.
    """
    renderers = renderers or default_report_renderers()
    artifacts: list[ReportRenderResult] = []
    diagnostics: list[Diagnostic] = []
    failed = 0
    for renderer in renderers:
        try:
            artifacts.append(renderer.render(snapshot, base_dir=base_dir))
        except Exception as exc:
            failed += 1
            diagnostics.append(
                Diagnostic(
                    "REPORT_RENDER_FAILED",
                    "A report renderer failed; committed results are unaffected.",
                    category=ErrorCategory.INTERNAL,
                    severity=DiagnosticSeverity.ERROR,
                    details=(
                        ("renderer", renderer.name),
                        ("error_type", type(exc).__name__),
                    ),
                )
            )
    value = ReportRenderOutcome(tuple(artifacts), len(artifacts), failed, snapshot.run_id)
    if failed == 0:
        outcome = OperationOutcome.COMPLETED
        counts = OperationCounts(succeeded=len(artifacts))
        return OperationResult(outcome, value, counts=counts, run_id=snapshot.run_id)
    if artifacts:
        outcome = OperationOutcome.PARTIAL
        counts = OperationCounts(succeeded=len(artifacts), failed=failed)
        return OperationResult(outcome, value, diagnostics=tuple(diagnostics), counts=counts, run_id=snapshot.run_id)
    outcome = OperationOutcome.FAILED
    counts = OperationCounts(failed=failed)
    return OperationResult(outcome, None, diagnostics=tuple(diagnostics), counts=counts, run_id=snapshot.run_id)


def _sha256(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _report_stem(snapshot: ReportSnapshot) -> str:
    return {
        "transbridge.polish-report.v1": "polish-report",
        "transbridge.mixed-report.v1": "mixed-report",
    }.get(snapshot.schema, "postprocess-report")


def _rotate_artifacts(base_dir: Path, stem: str, suffix: str, *, keep: int) -> None:
    artifacts = sorted(
        base_dir.glob(f"{stem}-*.{suffix}"),
        key=lambda path: (path.stat().st_mtime_ns, path.name),
        reverse=True,
    )
    for obsolete in artifacts[keep:]:
        obsolete.unlink(missing_ok=True)


def _csv_json(value: object) -> str:
    if value in (None, "", (), [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)


def _excel_value(value: object) -> object:
    if isinstance(value, str):
        return _spreadsheet_value(value)
    if value is None or isinstance(value, (int, float, bool)):
        return value
    serialized = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), allow_nan=False)
    return _spreadsheet_value(serialized)


def _spreadsheet_value(value: object) -> object:
    """Prevent untrusted translation text from becoming a spreadsheet formula."""
    if not isinstance(value, str) or not value:
        return value
    return f"'{value}" if value[0] in {"=", "+", "-", "@", "\t", "\r"} else value
