"""Canonical terminology model and local/ParaTranz format adapters."""

from __future__ import annotations

from collections.abc import Iterable, Mapping
import csv
from dataclasses import dataclass, field
import json
from pathlib import Path
import re
from typing import Any, Literal


@dataclass
class TermEntry:
    """Loss-minimizing terminology record used across all sources."""

    term: str
    translation: str
    source: str
    context: str = ""
    created_at: str = ""
    case_sensitive: bool = False
    variants: list[str] = field(default_factory=list)
    pos: str = ""
    note: str = ""
    external_id: int | str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


_ALIASES: dict[str, tuple[str, ...]] = {
    "term": ("term", "original", "source_term", "sourceTerm", "原文", "术语"),
    "translation": (
        "translation",
        "translated",
        "target_term",
        "targetTerm",
        "target",
        "译文",
        "翻译",
    ),
    "source": ("source", "entry_source", "term_source", "来源"),
    "context": ("context", "scope", "上下文", "语境"),
    "created_at": ("created_at", "createdAt", "创建时间"),
    "case_sensitive": ("case_sensitive", "caseSensitive", "大小写敏感"),
    "variants": ("variants", "variant", "变体"),
    "pos": ("pos", "part_of_speech", "partOfSpeech", "词性"),
    "note": ("note", "comment", "description", "注释", "备注"),
    "external_id": ("external_id", "id", "外部ID"),
    "metadata": ("metadata",),
}
_KNOWN_INPUT_KEYS = {alias for aliases in _ALIASES.values() for alias in aliases}
_KNOWN_INPUT_KEYS_FOLDED = {key.casefold() for key in _KNOWN_INPUT_KEYS}
_TERM_HEADERS = frozenset({*(alias.casefold() for alias in _ALIASES["term"]), "source"})
_TRANSLATION_HEADERS = frozenset(alias.casefold() for alias in _ALIASES["translation"])
_VARIANT_SPLIT_RE = re.compile(r"[|;,，；\n]+")


def term_entry_from_mapping(item: Mapping[str, Any], *, source: str | None = None) -> TermEntry | None:
    """Normalize one mapping, retaining unrecognized fields in ``metadata``."""

    term = _text(_first(item, "term"))
    translation = _text(_first(item, "translation"))
    if not term or not translation:
        return None

    metadata = _metadata(_first(item, "metadata"))
    metadata.update({
        str(key): _json_safe(value)
        for key, value in item.items()
        if str(key).strip().casefold() not in _KNOWN_INPUT_KEYS_FOLDED
    })

    configured_source = source if source is not None else _text(_first(item, "source"))
    return TermEntry(
        term=term,
        translation=translation,
        source=configured_source,
        context=_text(_first(item, "context")),
        created_at=_text(_first(item, "created_at")),
        case_sensitive=_bool(_first(item, "case_sensitive")),
        variants=_variants(_first(item, "variants")),
        pos=_text(_first(item, "pos")),
        note=_text(_first(item, "note")),
        external_id=_identifier(_first(item, "external_id")),
        metadata=metadata,
    )


def term_entries_from_data(data: Any, *, source: str | None) -> list[TermEntry]:
    """Normalize supported JSON-shaped containers into canonical entries."""

    records: Any = data
    if isinstance(data, Mapping):
        wrapped = next(
            (data[key] for key in ("terms", "results", "items", "data") if isinstance(data.get(key), list)),
            None,
        )
        if wrapped is not None:
            records = wrapped
        elif _looks_like_term_record(data):
            records = [data]
        else:
            records = [
                (
                    {"term": term, **dict(translation)}
                    if isinstance(translation, Mapping)
                    else {
                        "term": term,
                        "translation": translation,
                    }
                )
                for term, translation in data.items()
            ]

    if not isinstance(records, list):
        raise ValueError("术语数据必须是数组、术语对象、键值映射或 terms/results 包装对象")

    entries: list[TermEntry] = []
    for record in records:
        if not isinstance(record, Mapping):
            continue
        entry = term_entry_from_mapping(record, source=source)
        if entry is not None:
            entries.append(entry)
    return entries


def term_entry_to_canonical_dict(entry: TermEntry) -> dict[str, Any]:
    """Serialize every canonical field without leaking metadata into reserved keys."""

    return {
        "term": entry.term,
        "translation": entry.translation,
        "source": entry.source,
        "context": entry.context,
        "created_at": entry.created_at,
        "case_sensitive": entry.case_sensitive,
        "variants": list(entry.variants),
        "pos": entry.pos,
        "note": entry.note,
        "external_id": entry.external_id,
        "metadata": _json_safe(entry.metadata),
    }


def term_entry_to_paratranz_dict(entry: TermEntry) -> dict[str, Any]:
    """Create a ParaTranz writable payload and omit server-owned fields."""

    payload: dict[str, Any] = {
        "term": entry.term,
        "translation": entry.translation,
        "variants": list(entry.variants),
        "caseSensitive": entry.case_sensitive,
    }
    if entry.pos:
        payload["pos"] = entry.pos
    if entry.note:
        payload["note"] = entry.note
    return payload


def load_terms_json(path: str | Path, *, source: str | None = "json") -> list[TermEntry]:
    try:
        with Path(path).open(encoding="utf-8-sig") as stream:
            return term_entries_from_data(json.load(stream), source=source)
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        raise ValueError(f"无法读取 JSON 术语库 {path}: {exc}") from exc


def dump_terms_json(
    path: str | Path,
    entries: Iterable[TermEntry],
    *,
    target: Literal["canonical", "paratranz"] = "canonical",
) -> None:
    if target not in {"canonical", "paratranz"}:
        raise ValueError(f"未知的术语 JSON 目标格式: {target}")
    serializer = term_entry_to_paratranz_dict if target == "paratranz" else term_entry_to_canonical_dict
    with Path(path).open("w", encoding="utf-8") as stream:
        json.dump([serializer(entry) for entry in entries], stream, ensure_ascii=False, indent=2)


def load_terms_csv(path: str | Path, *, source: str | None = "csv") -> list[TermEntry]:
    try:
        with Path(path).open(encoding="utf-8-sig", newline="") as stream:
            sample = stream.read(4096)
            stream.seek(0)
            dialect = _csv_dialect(sample)
            reader = csv.reader(stream, dialect)
            first_row = next(reader, None)
            if first_row is None:
                return []
            entries: list[TermEntry] = []
            if _has_term_headers(first_row):
                headers = [_text(value) for value in first_row]
                records = (dict(zip(headers, row, strict=False)) for row in reader)
            else:
                records = (
                    {"term": row[0], "translation": row[1]} for row in _chain_first(first_row, reader) if len(row) >= 2
                )
            for record in records:
                entry = term_entry_from_mapping(record, source=source)
                if entry is not None:
                    entries.append(entry)
            return entries
    except (OSError, csv.Error) as exc:
        raise ValueError(f"无法读取 CSV 术语库 {path}: {exc}") from exc


def dump_terms_csv(path: str | Path, entries: Iterable[TermEntry]) -> None:
    fieldnames = [
        "term",
        "translation",
        "pos",
        "note",
        "variants",
        "case_sensitive",
        "context",
        "created_at",
        "external_id",
        "source",
        "metadata",
    ]
    with Path(path).open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fieldnames)
        writer.writeheader()
        for entry in entries:
            row = term_entry_to_canonical_dict(entry)
            row["variants"] = json.dumps(row["variants"], ensure_ascii=False)
            row["metadata"] = json.dumps(row["metadata"], ensure_ascii=False, sort_keys=True)
            writer.writerow(row)


def load_terms_excel(
    path: str | Path,
    *,
    source: str | None = "excel",
    original_column: str = "A",
    translation_column: str = "B",
) -> list[TermEntry]:
    if Path(path).suffix.casefold() == ".xls":
        return _load_terms_xls(
            path,
            source=source,
            original_column=original_column,
            translation_column=translation_column,
        )
    workbook = None
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        return _entries_from_rows(
            sheet.iter_rows(values_only=True),
            source=source,
            original_column=original_column,
            translation_column=translation_column,
        )
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 术语库 {path}: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.close()


def _load_terms_xls(
    path: str | Path,
    *,
    source: str | None,
    original_column: str,
    translation_column: str,
) -> list[TermEntry]:
    workbook = None
    try:
        import xlrd

        workbook = xlrd.open_workbook(path, on_demand=True)
        if workbook.nsheets == 0:
            return []
        sheet = workbook.sheet_by_index(0)
        rows = (sheet.row_values(index) for index in range(sheet.nrows))
        return _entries_from_rows(
            rows,
            source=source,
            original_column=original_column,
            translation_column=translation_column,
        )
    except Exception as exc:
        raise ValueError(f"无法读取旧式 Excel 术语库 {path}: {exc}") from exc
    finally:
        if workbook is not None:
            workbook.release_resources()


def _entries_from_rows(
    rows: Iterable[Iterable[Any]],
    *,
    source: str | None,
    original_column: str,
    translation_column: str,
) -> list[TermEntry]:
    iterator = iter(rows)
    first_row = list(next(iterator, ()))
    if not first_row:
        return []
    headers = [_text(value) for value in first_row]
    if _has_term_headers(headers):
        records = (dict(zip(headers, row, strict=False)) for row in iterator)
    else:
        original_index = column_letter_to_index(original_column)
        translation_index = column_letter_to_index(translation_column)
        records = (
            {
                "term": row[original_index] if original_index < len(row) else None,
                "translation": row[translation_index] if translation_index < len(row) else None,
            }
            for row_values in iterator
            if (row := list(row_values))
        )
    entries: list[TermEntry] = []
    for record in records:
        entry = term_entry_from_mapping(record, source=source)
        if entry is not None:
            entries.append(entry)
    return entries


def dump_terms_excel(path: str | Path, entries: Iterable[TermEntry]) -> None:
    import openpyxl

    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet()
    headers = [
        "term",
        "translation",
        "pos",
        "note",
        "variants",
        "case_sensitive",
        "context",
        "created_at",
        "external_id",
        "source",
        "metadata",
    ]
    sheet.append(headers)
    for entry in entries:
        row = term_entry_to_canonical_dict(entry)
        row["variants"] = json.dumps(row["variants"], ensure_ascii=False)
        row["metadata"] = json.dumps(row["metadata"], ensure_ascii=False, sort_keys=True)
        sheet.append([row[header] for header in headers])
    workbook.save(path)


def column_letter_to_index(letter: str) -> int:
    normalized = letter.upper().strip()
    if not normalized or any(character < "A" or character > "Z" for character in normalized):
        raise ValueError(f"无效的 Excel 列名: {letter!r}")
    result = 0
    for character in normalized:
        result = result * 26 + ord(character) - ord("A") + 1
    return result - 1


def _first(item: Mapping[str, Any], field_name: str) -> Any:
    for alias in _ALIASES[field_name]:
        if alias in item and item[alias] is not None:
            return item[alias]
    folded = {str(key).strip().casefold(): value for key, value in item.items()}
    for alias in _ALIASES[field_name]:
        if alias.casefold() in folded and folded[alias.casefold()] is not None:
            return folded[alias.casefold()]
    if (
        field_name == "term"
        and "source" in folded
        and any(alias.casefold() in folded for alias in _ALIASES["translation"] if alias != "translation")
    ):
        return folded["source"]
    return None


def _looks_like_term_record(data: Mapping[str, Any]) -> bool:
    keys = {str(key).strip().casefold() for key in data}
    return bool(keys & _TERM_HEADERS) and bool(keys & _TRANSLATION_HEADERS)


def _has_term_headers(row: Iterable[Any]) -> bool:
    headers = {_text(value).casefold() for value in row}
    return bool(headers & _TERM_HEADERS) and bool(headers & _TRANSLATION_HEADERS)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    return _text(value).casefold() in {"1", "true", "yes", "on", "是", "y"}


def _identifier(value: Any) -> int | str | None:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = _text(value)
    try:
        return int(text)
    except ValueError:
        return text


def _metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, str) and value.strip():
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {"value": value}
        if isinstance(parsed, Mapping):
            return {str(key): _json_safe(item) for key, item in parsed.items()}
    return {}


def _variants(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
        except json.JSONDecodeError:
            values: Iterable[Any] = _VARIANT_SPLIT_RE.split(text)
        else:
            values = parsed if isinstance(parsed, list) else [parsed]
    elif isinstance(value, Iterable) and not isinstance(value, Mapping):
        values = value
    else:
        values = [value]
    result: list[str] = []
    for item in values:
        normalized = _text(item)
        if normalized and normalized not in result:
            result.append(normalized)
    return result


def _json_safe(value: Any) -> Any:
    try:
        json.dumps(value, ensure_ascii=False, allow_nan=False)
        return value
    except (TypeError, ValueError):
        if isinstance(value, Mapping):
            return {str(key): _json_safe(item) for key, item in value.items()}
        if isinstance(value, Iterable) and not isinstance(value, (str, bytes)):
            return [_json_safe(item) for item in value]
        return str(value)


def _csv_dialect(sample: str) -> type[csv.Dialect] | csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return csv.excel


def _chain_first(first: list[str], remaining: Iterable[list[str]]) -> Iterable[list[str]]:
    yield first
    yield from remaining


__all__ = [
    "TermEntry",
    "column_letter_to_index",
    "dump_terms_csv",
    "dump_terms_excel",
    "dump_terms_json",
    "load_terms_csv",
    "load_terms_excel",
    "load_terms_json",
    "term_entries_from_data",
    "term_entry_from_mapping",
    "term_entry_to_canonical_dict",
    "term_entry_to_paratranz_dict",
]
