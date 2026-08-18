"""AI翻译/润色结果报告生成器。

生成结构化 Excel 报告（.xlsx），支持翻译模式（5 Sheet）和润色模式（3 Sheet）。
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

if TYPE_CHECKING:
    from transbridge.ai_translator.translator import TranslationResult
    from transbridge.ai_translator.post_processor.base import PostProcessResult
    from transbridge.ai_translator.post_processor.polisher import PolishResult
    from transbridge.converter.translation_entry import TranslationEntry

_logger = logging.getLogger(__name__)

# ── 样式常量 ─────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CELL_FONT = Font(name="微软雅黑", size=10)
_WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


class ReportGenerator:
    """报告生成器，负责聚合翻译/后处理/润色结果数据并生成 Excel 文件。

    用法:
        generator = ReportGenerator("MyPlugin")
        path = generator.generate_translate_report(result, refine, polish, decisions)
        # 或
        path = generator.generate_polish_report(polish_results, entries, stats)
    """

    def __init__(self, esp_stem: str):
        self._esp_stem = esp_stem
        from transbridge.paratranz.config_manager import LLMConfig
        ai_dir = LLMConfig.get_ai_translator_dir(esp_stem)
        self._output_dir = os.path.join(ai_dir, "reports")

    # ── 公共方法 ──────────────────────────────────────────────────────────────

    def generate_translate_report(
        self,
        result: "TranslationResult",
        refine_results: dict | None = None,
        polish_results: dict | None = None,
        decisions: dict | None = None,
    ) -> str | None:
        """生成翻译模式报告（5 Sheet Excel）。

        Args:
            result: 翻译结果
            refine_results: {entry_id: RefineResult}，修复中间数据
            polish_results: {entry_id: PolishResult}，润色中间数据
            decisions: {entry_id: ArbiterDecision}，裁决中间数据

        Returns:
            生成的报告文件绝对路径，失败返回 None
        """
        pp = result.post_process_result
        refine = refine_results or (pp.refine_results if pp else None) or {}
        polish = polish_results or (pp.polish_results if pp else None) or {}
        decs = decisions or (pp.decisions if pp else None) or {}

        try:
            os.makedirs(self._output_dir, exist_ok=True)
            wb = Workbook()

            # Sheet 1: Summary
            self._write_translate_summary(wb, result, pp, refine, polish, decs)

            # Sheet 2: Entries
            self._write_translate_entries(wb, result, pp, refine, polish, decs)

            # Sheet 3: Issues
            self._write_issues(wb, pp)

            # Sheet 4: Refinements
            self._write_refinements(wb, refine)

            # Sheet 5: Arbitrations
            self._write_arbitrations(wb, decs)

            # 删除默认 sheet
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            # 保存
            filepath = self._save(wb, "translate")
            self._rotate()
            return filepath

        except Exception:
            _logger.exception("翻译报告生成失败")
            return None

    def generate_polish_report(
        self,
        polish_results: dict[str, "PolishResult"],
        entries: list["TranslationEntry"],
        stats: dict,
    ) -> str | None:
        """生成润色模式报告（3 Sheet Excel）。

        Args:
            polish_results: {entry_id: PolishResult}
            entries: 被润色的条目列表
            stats: {"total": int, "accepted": int, "rejected": int,
                    "failed": int, "polish_level": str, "avg_confidence": float}

        Returns:
            生成的报告文件绝对路径，失败返回 None
        """
        try:
            os.makedirs(self._output_dir, exist_ok=True)
            wb = Workbook()

            # Sheet 1: Summary
            self._write_polish_summary(wb, stats)

            # Sheet 2: Entries
            self._write_polish_entries(wb, entries, polish_results)

            # Sheet 3: Polish Details
            self._write_polish_details(wb, polish_results)

            # 删除默认 sheet
            if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
                del wb["Sheet"]

            # 保存
            filepath = self._save(wb, "polish")
            self._rotate()
            return filepath

        except Exception:
            _logger.exception("润色报告生成失败")
            return None

    # ── 翻译报告 Sheet 写入 ───────────────────────────────────────────────────

    def _write_translate_summary(self, wb, result, pp, refine, polish, decs):
        ws = wb.active
        ws.title = "Summary"

        error_count = 0
        warning_count = 0
        info_count = 0
        if pp:
            for i in pp.issues:
                if i.severity == "error":
                    error_count += 1
                elif i.severity == "warning":
                    warning_count += 1
                elif i.severity == "info":
                    info_count += 1

        passed = sum(1 for d in decs.values() if getattr(d, "verdict", None) == "pass")
        rejected = sum(1 for d in decs.values() if getattr(d, "verdict", None) == "reject")
        pending = sum(1 for d in decs.values() if getattr(d, "verdict", None) == "pending")

        rows = [
            ["指标", "值", "说明"],
            ["total_checked", pp.total_checked if pp else 0, "检查后处理的总条目数"],
            ["success_count", result.success_count, "翻译成功条目数"],
            ["failed_count", result.failed_count, "翻译失败条目数"],
            ["skipped_count", result.skipped_count, "翻译跳过条目数"],
            ["new_dynamic_terms", result.new_dynamic_terms, "新增动态术语数"],
            ["issue_count", pp.issue_count if pp else 0, "发现问题总数"],
            ["error_count", error_count, "error 级别问题数"],
            ["warning_count", warning_count, "warning 级别问题数"],
            ["info_count", info_count, "info 级别问题数"],
            ["passed", passed, "裁决通过条目数"],
            ["rejected", rejected, "裁决打回条目数"],
            ["pending", pending, "裁决待审条目数"],
            ["refined_count", len(refine), "经历 LLM 修复的条目数"],
            ["polished_count", len(polish), "经历 LLM 润色的条目数"],
            ["timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "报告生成时间"],
            ["esp_stem", self._esp_stem, "来源插件名"],
        ]

        self._write_rows(ws, rows)

    def _write_translate_entries(self, wb, result, pp, refine, polish, decs):
        ws = wb.create_sheet("Entries")

        headers = [
            "entry_id", "original", "initial_translation",
            "refined_translation", "polished_translation", "final_translation",
            "stage", "verdict", "verdict_reason", "confidence",
            "issue_count", "issue_types",
        ]
        self._write_header(ws, headers)

        # 需要从 decisions 反向索引 entry 数据
        # decisions 的 key 是 entry_id，我们通过 pp 和中间数据重建条目视图
        all_entry_ids = set()
        all_entry_ids.update(decs.keys())
        all_entry_ids.update(refine.keys())
        all_entry_ids.update(polish.keys())
        if pp:
            all_entry_ids.update(i.entry_id for i in pp.issues)

        row_idx = 2
        for entry_id in sorted(all_entry_ids):
            dec = decs.get(entry_id)
            ref = refine.get(entry_id)
            pol = polish.get(entry_id)

            # 确定最终译文
            if pol and pol.polished_translation:
                final = pol.polished_translation
            elif ref and ref.refined_translation:
                final = ref.refined_translation
            else:
                final = ref.original_translation if ref else ""

            # 收集问题
            entry_issues = []
            if pp:
                entry_issues = [i for i in pp.issues if i.entry_id == entry_id]

            stage = 1 if (dec and dec.verdict == "pass") else (0 if (dec and dec.verdict == "reject") else 2)

            row = [
                entry_id,
                entry_issues[0].original if entry_issues else "",
                ref.original_translation if ref else (pol.original_translation if pol else ""),
                ref.refined_translation if ref else "",
                pol.polished_translation if pol else "",
                final,
                stage,
                dec.verdict if dec else "",
                dec.reason if dec else "",
                f"{dec.confidence:.0%}" if dec else "",
                len(entry_issues),
                ", ".join(i.issue_type for i in entry_issues),
            ]
            self._write_row(ws, row_idx, row)
            row_idx += 1

        self._auto_width(ws)

    def _write_issues(self, wb, pp):
        ws = wb.create_sheet("Issues")
        headers = ["entry_id", "issue_type", "severity", "message", "suggestion", "original", "translation"]
        self._write_header(ws, headers)

        if pp and pp.issues:
            for row_idx, issue in enumerate(pp.issues, start=2):
                row = [
                    issue.entry_id, issue.issue_type, issue.severity,
                    issue.message, issue.suggestion, issue.original, issue.translation,
                ]
                self._write_row(ws, row_idx, row)

        self._auto_width(ws)

    def _write_refinements(self, wb, refine):
        ws = wb.create_sheet("Refinements")
        headers = ["entry_id", "refined_translation", "confidence", "fixes_applied", "note"]
        self._write_header(ws, headers)

        for row_idx, (entry_id, ref) in enumerate(sorted(refine.items()), start=2):
            fixes_text = ""
            if hasattr(ref, "fixes_applied") and ref.fixes_applied:
                fixes_text = "; ".join(
                    getattr(f, "description", str(f)) for f in ref.fixes_applied
                )
            row = [
                entry_id, ref.refined_translation,
                f"{ref.confidence:.0%}", fixes_text,
                getattr(ref, "note", ""),
            ]
            self._write_row(ws, row_idx, row)

        self._auto_width(ws)

    def _write_arbitrations(self, wb, decs):
        ws = wb.create_sheet("Arbitrations")
        headers = ["entry_id", "verdict", "reason", "confidence", "suggested_action"]
        self._write_header(ws, headers)

        for row_idx, (entry_id, dec) in enumerate(sorted(decs.items()), start=2):
            row = [
                entry_id, dec.verdict, dec.reason,
                f"{dec.confidence:.0%}", dec.suggested_action,
            ]
            self._write_row(ws, row_idx, row)

        self._auto_width(ws)

    # ── 润色报告 Sheet 写入 ───────────────────────────────────────────────────

    def _write_polish_summary(self, wb, stats):
        ws = wb.active
        ws.title = "Summary"

        rows = [
            ["指标", "值", "说明"],
            ["total_entries", stats.get("total", 0), "润色条目总数"],
            ["accepted_count", stats.get("accepted", 0), "用户接受数"],
            ["rejected_count", stats.get("rejected", 0), "用户拒绝数"],
            ["failed_count", stats.get("failed", 0), "润色失败数"],
            ["polish_level", stats.get("polish_level", ""), "润色强度 (light/moderate/aggressive)"],
            ["avg_confidence", f"{stats.get('avg_confidence', 0):.1%}", "平均信心度"],
            ["timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "报告生成时间"],
            ["esp_stem", self._esp_stem, "来源插件名"],
        ]

        self._write_rows(ws, rows)

    def _write_polish_entries(self, wb, entries, polish_results):
        ws = wb.create_sheet("Entries")
        headers = [
            "entry_id", "original", "original_translation",
            "polished_translation", "accepted", "confidence", "changes_summary",
        ]
        self._write_header(ws, headers)

        for row_idx, entry in enumerate(entries, start=2):
            pr = polish_results.get(entry.id)
            if pr:
                changes_summary = "; ".join(
                    c.get("aspect", "") + ": " + c.get("before", "") + " → " + c.get("after", "")
                    for c in pr.changes
                ) if pr.changes else ""
                accepted = pr.confidence > 0.0  # 成功即为接受
                row = [
                    entry.id, entry.original, pr.original_translation,
                    pr.polished_translation,
                    "是" if accepted else "否",
                    f"{pr.confidence:.0%}",
                    changes_summary[:500] if changes_summary else "",
                ]
            else:
                row = [entry.id, entry.original, entry.translation or "", "", "否", "0%", ""]
            self._write_row(ws, row_idx, row)

        self._auto_width(ws)

    def _write_polish_details(self, wb, polish_results):
        ws = wb.create_sheet("Polish")
        headers = ["entry_id", "change_aspect", "before", "after", "reason"]
        self._write_header(ws, headers)

        row_idx = 2
        for entry_id, pr in sorted(polish_results.items()):
            if pr.changes:
                for change in pr.changes:
                    row = [
                        entry_id,
                        change.get("aspect", ""),
                        change.get("before", ""),
                        change.get("after", ""),
                        change.get("reason", ""),
                    ]
                    self._write_row(ws, row_idx, row)
                    row_idx += 1

        self._auto_width(ws)

    # ── 文件管理 ──────────────────────────────────────────────────────────────

    def _save(self, wb: Workbook, mode: str) -> str:
        """保存工作簿并返回文件路径。"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # 清理文件名中的非法字符
        safe_stem = re.sub(r'[<>:"/\\|?*]', '_', self._esp_stem)
        filename = f"{safe_stem}_{mode}_report_{timestamp}.xlsx"
        filepath = os.path.join(self._output_dir, filename)
        wb.save(filepath)
        _logger.info("报告已保存: %s", filepath)
        return filepath

    def _rotate(self, keep: int = 20) -> None:
        """清理旧报告，保留最近 N 份。"""
        try:
            if not os.path.isdir(self._output_dir):
                return
            files = [
                os.path.join(self._output_dir, f)
                for f in os.listdir(self._output_dir)
                if f.endswith(".xlsx")
            ]
            if len(files) <= keep:
                return
            # 按修改时间排序，旧的在前面
            files.sort(key=lambda p: os.path.getmtime(p))
            to_delete = files[: len(files) - keep]
            for f in to_delete:
                try:
                    os.remove(f)
                    _logger.info("已清理旧报告: %s", os.path.basename(f))
                except OSError:
                    _logger.warning("无法删除旧报告: %s", f)
        except Exception:
            _logger.exception("报告清理异常")

    # ── 工具方法 ──────────────────────────────────────────────────────────────

    @staticmethod
    def _write_header(ws, headers: list[str], row: int = 1):
        """写入表头行。"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _write_row(ws, row: int, values: list):
        """写入数据行。"""
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = _CELL_FONT
            cell.border = _THIN_BORDER
            if isinstance(value, str) and len(value) > 50:
                cell.alignment = _WRAP_ALIGNMENT

    @staticmethod
    def _write_rows(ws, rows: list[list]):
        """批量写入行（含表头）。"""
        for row_idx, row_data in enumerate(rows, start=1):
            if row_idx == 1:
                ReportGenerator._write_header(ws, row_data)
            else:
                ReportGenerator._write_row(ws, row_idx, row_data)
        ReportGenerator._auto_width(ws)

    @staticmethod
    def _auto_width(ws):
        """自动调整列宽（基于内容估算）。"""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if col_letter is None and cell.column_letter:
                    col_letter = cell.column_letter
                if cell.value:
                    # 中文字符按2个字符宽度估算
                    text = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in text)
                    max_len = max(max_len, length)
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
