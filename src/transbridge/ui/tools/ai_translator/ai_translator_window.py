"""
AI 翻译配置窗口。

AITranslatorWindow  — 配置窗口，翻译开始前使用
进度窗口见 _translation_progress_window.py
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QScrollArea,
    QLineEdit, QComboBox, QSpinBox, QPushButton,
    QRadioButton, QButtonGroup, QFileDialog, QMessageBox,
    QStackedWidget,
    QCheckBox, QListWidget, QListWidgetItem,
    QAbstractItemView, QFrame, QTabWidget,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor, QBrush

from src.transbridge.ui.tools.ai_translator._translation_worker import _TranslationWorker
from src.transbridge.ui.tools.ai_translator._translation_progress_window import _TranslationProgressWindow
from src.transbridge.ui.tools.ai_translator._term_editor_dialog import _TermEditorDialog
from src.transbridge.paratranz.config_manager import apply_rules

if TYPE_CHECKING:
    from src.transbridge.ui.context import AppContext
    from src.transbridge.ui.workbench.step2 import Step2PreviewWidget


class AITranslatorWindow(QWidget):
    """AI 翻译配置窗口。点击「开始翻译」后关闭自身并弹出进度窗口。"""

    # 翻译启动后发出，携带进度窗口实例
    progress_window_created = pyqtSignal(object)

    def __init__(self, ctx: "AppContext", step2: "Step2PreviewWidget", parent=None):
        super().__init__(parent, Qt.WindowType.Window)
        self._ctx = ctx
        self._step2 = step2
        self.setWindowTitle("AI 自动翻译")
        self.resize(680, 520)
        self._init_ui()
        self._load_config()
        self._connect_auto_save()
        self._check_checkpoint()

    @classmethod
    def open_for_translation(
        cls,
        ctx: "AppContext",
        step2: "Step2PreviewWidget",
        parent=None,
    ) -> QWidget | None:
        """打开翻译入口：先选择目标，再进入相应流程。

        Returns:
            打开的窗口实例，或 None（用户取消）
        """
        # 检查是否有已加载的插件
        if not ctx.slots:
            QMessageBox.warning(parent, "AI 翻译", "请先加载插件。")
            return None

        # 延迟导入批量翻译相关模块
        from src.transbridge.ui.tools.ai_translator._translation_target_dialog import _TranslationTargetDialog

        # 弹出目标选择对话框
        target_dlg = _TranslationTargetDialog(ctx, parent)
        if target_dlg.exec() != target_dlg.DialogCode.Accepted:
            return None

        if target_dlg.is_batch_mode():
            # 批量翻译模式
            return cls._open_batch_mode(ctx, parent)
        else:
            # 单插件模式
            window = cls(ctx, step2, parent)
            window.show()
            return window

    @classmethod
    def _open_batch_mode(cls, ctx: "AppContext", parent=None) -> QWidget | None:
        """打开批量翻译流程。"""
        # 延迟导入批量翻译相关模块
        from src.transbridge.ui.tools.ai_translator._batch_translation_dialog import _BatchTranslationDialog
        from src.transbridge.ui.tools.ai_translator._batch_translation_worker import _BatchTranslationWorker
        from src.transbridge.ui.tools.ai_translator._batch_translation_progress_window import _BatchTranslationProgressWindow

        # 弹出批量翻译对话框（包含配置编辑）
        batch_dlg = _BatchTranslationDialog(ctx, parent)
        if batch_dlg.exec() != batch_dlg.DialogCode.Accepted:
            return None

        selected_slots = batch_dlg.get_selected_slots()
        if not selected_slots:
            QMessageBox.warning(parent, "批量翻译", "请至少选择一个插件。")
            return None

        overwrite = batch_dlg.is_overwrite()
        llm_config = batch_dlg.get_llm_config()

        if not llm_config or not llm_config.api_key:
            QMessageBox.warning(parent, "批量翻译", "请先配置 API Key。")
            return None
        if not llm_config.model:
            QMessageBox.warning(parent, "批量翻译", "请先配置模型名。")
            return None

        # 检查术语库（可选警告）
        if not ctx.current_project and cls._check_all_terms_empty_batch(llm_config, selected_slots):
            reply = QMessageBox.question(
                parent, "术语库为空",
                "当前未选择 ParaTranz 项目，且所有术语来源均为空。\n\n"
                "没有术语库辅助，翻译质量可能下降。是否继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return None

        # 创建 ParaTranz 客户端（如果选择了项目）
        paratranz_client = None
        project_id = None
        if ctx.current_project:
            project_id = ctx.current_project.get("id")
            from src.transbridge.paratranz.api.paratranz_terms_api import ParatranzTermsAPI
            paratranz_client = ParatranzTermsAPI(ctx.config)

        # 创建 Worker 和进度窗口
        worker = _BatchTranslationWorker(
            slots=selected_slots,
            llm_config=llm_config,
            overwrite=overwrite,
            paratranz_client=paratranz_client,
            project_id=project_id,
        )
        progress_win = _BatchTranslationProgressWindow(worker, ctx)

        progress_win.show()
        worker.start()

        return progress_win

    @classmethod
    def _check_all_terms_empty_batch(cls, cfg, slots: list) -> bool:
        """检查所有术语来源是否均为空。"""
        import os
        from src.transbridge.ai_translator.term_database import DynamicTermDatabase

        # 检查第一个插件的动态术语库
        if slots:
            esp_path = slots[0].esp_path
            if esp_path:
                dynamic_db = DynamicTermDatabase(esp_path)
                dynamic_db.load()
                if dynamic_db.as_list():
                    return False

        if cfg.local_json_path and os.path.exists(cfg.local_json_path):
            try:
                import json
                with open(cfg.local_json_path, encoding="utf-8") as f:
                    data = json.load(f)
                if data:
                    return False
            except Exception:
                pass

        if cfg.local_excel_path and os.path.exists(cfg.local_excel_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(cfg.local_excel_path, read_only=True, data_only=True)
                ws = wb.active
                col_orig = cls._col_letter_to_index(cfg.excel_original_col or "A")
                col_trans = cls._col_letter_to_index(cfg.excel_translation_col or "B")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        term = row[col_orig] if col_orig < len(row) else None
                        trans = row[col_trans] if col_trans < len(row) else None
                        if term and trans:
                            return False
                    except (IndexError, TypeError):
                        continue
            except Exception:
                pass

        return True

    # ── UI 构建 ───────────────────────────────────────────────────────────────

    def _init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(8)

        # ── 模式切换 ────────────────────────────────────────────────────────────
        mode_box = QHBoxLayout()
        mode_box.addWidget(QLabel("模式:"))
        self._mode_group = QButtonGroup(self)
        self._mode_translate = QRadioButton("翻译")
        self._mode_polish = QRadioButton("润色")
        self._mode_mixed = QRadioButton("混合")
        self._mode_group.addButton(self._mode_translate)
        self._mode_group.addButton(self._mode_polish)
        self._mode_group.addButton(self._mode_mixed)
        self._mode_translate.setChecked(True)
        mode_box.addWidget(self._mode_translate)
        mode_box.addWidget(self._mode_polish)
        mode_box.addWidget(self._mode_mixed)
        mode_box.addStretch()
        main_layout.addLayout(mode_box)

        # ── LLM 配置区 ────────────────────────────────────────────────────────
        llm_box = QGroupBox("LLM 配置")
        llm_layout = QVBoxLayout(llm_box)
        llm_layout.setSpacing(4)

        def _row(label_text, widget):
            row = QHBoxLayout()
            lbl = QLabel(label_text)
            lbl.setFixedWidth(90)
            row.addWidget(lbl)
            row.addWidget(widget)
            return row

        self._provider_combo = QComboBox()
        self._provider_combo.addItems(["OpenAI 兼容", "Anthropic"])
        self._provider_combo.currentIndexChanged.connect(self._on_provider_changed)
        llm_layout.addLayout(_row("供应商:", self._provider_combo))

        self._target_lang_combo = QComboBox()
        self._target_lang_combo.addItems(["zh_CN"])
        self._target_lang_combo.setToolTip("目标语言配置，对应 data/prompts/langs/{lang}.toml")
        llm_layout.addLayout(_row("目标语言:", self._target_lang_combo))

        self._model_edit = QLineEdit()
        self._model_edit.setPlaceholderText("如 gpt-4o / deepseek-chat")
        llm_layout.addLayout(_row("模型名:", self._model_edit))

        self._apikey_edit = QLineEdit()
        self._apikey_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self._apikey_edit.setPlaceholderText("API Key")
        llm_layout.addLayout(_row("API Key:", self._apikey_edit))

        self._baseurl_edit = QLineEdit()
        self._baseurl_edit.setPlaceholderText("https://api.openai.com/v1")
        llm_layout.addLayout(_row("Base URL:", self._baseurl_edit))

        self._concurrent_spin = QSpinBox()
        self._concurrent_spin.setRange(1, 50)
        self._concurrent_spin.setValue(20)
        llm_layout.addLayout(_row("并发数:", self._concurrent_spin))

        self._tokens_spin = QSpinBox()
        self._tokens_spin.setRange(200, 32000)
        self._tokens_spin.setSingleStep(200)
        self._tokens_spin.setValue(2500)
        llm_layout.addLayout(_row("拆批 Token:", self._tokens_spin))

        self._output_tokens_spin = QSpinBox()
        self._output_tokens_spin.setRange(0, 65536)
        self._output_tokens_spin.setSingleStep(256)
        self._output_tokens_spin.setValue(0)
        self._output_tokens_spin.setSpecialValueText("不限制（模型默认）")
        llm_layout.addLayout(_row("输出 Token:", self._output_tokens_spin))

        self._max_terms_spin = QSpinBox()
        self._max_terms_spin.setRange(10, 500)
        self._max_terms_spin.setValue(50)
        self._max_terms_spin.setToolTip("每批次发送给 LLM 的术语表上限，防止 token 超限")
        llm_layout.addLayout(_row("术语上限:", self._max_terms_spin))

        test_btn = QPushButton("测试连接")
        test_btn.setFixedWidth(100)
        test_btn.clicked.connect(self._on_test_connection)
        test_row = QHBoxLayout()
        test_row.addStretch()
        test_row.addWidget(test_btn)
        llm_layout.addLayout(test_row)

        # 创建标签页
        self._tabs = QTabWidget()

        # Tab 1: LLM 与模型
        tab_llm = QWidget()
        tab_llm_layout = QVBoxLayout(tab_llm)
        tab_llm_layout.setSpacing(6)
        tab_llm_layout.addWidget(llm_box)

        # ── Embedding 配置区 ──────────────────────────────────────────────────
        embed_box = QGroupBox("语义检索配置（Embedding）")
        embed_layout = QVBoxLayout(embed_box)
        embed_layout.setSpacing(4)

        self._embed_provider_combo = QComboBox()
        self._embed_provider_combo.addItems(["本地模型（sentence-transformers）", "API 服务（OpenAI 兼容）"])
        self._embed_provider_combo.setToolTip(
            "本地模型：使用 sentence-transformers，无需网络，需要安装依赖\n"
            "API 服务：使用 OpenAI/DeepSeek/阿里云等兼容 embedding 接口"
        )
        self._embed_provider_combo.currentIndexChanged.connect(self._on_embed_provider_changed)
        embed_layout.addLayout(_row("模式:", self._embed_provider_combo))

        # 本地模型配置
        self._embed_local_model_label = QLabel("模型名:")
        self._embed_local_model_label.setFixedWidth(90)
        self._embed_local_model_edit = QLineEdit()
        self._embed_local_model_edit.setPlaceholderText("paraphrase-multilingual-MiniLM-L12-v2")
        self._embed_local_model_edit.setToolTip("本地模型名称，首次使用会从 HuggingFace 下载")
        local_row = QHBoxLayout()
        local_row.addWidget(self._embed_local_model_label)
        local_row.addWidget(self._embed_local_model_edit)
        embed_layout.addLayout(local_row)

        # API 配置（OpenAI 兼容）
        self._embed_model_label = QLabel("API 模型:")
        self._embed_model_label.setFixedWidth(90)
        self._embed_model_edit = QLineEdit()
        self._embed_model_edit.setPlaceholderText("text-embedding-3-small")
        self._embed_model_edit.setToolTip("API embedding 模型名称")
        model_row = QHBoxLayout()
        model_row.addWidget(self._embed_model_label)
        model_row.addWidget(self._embed_model_edit)
        embed_layout.addLayout(model_row)

        self._embed_apikey_label = QLabel("API Key:")
        self._embed_apikey_label.setFixedWidth(90)
        self._embed_apikey_edit = QLineEdit()
        self._embed_apikey_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self._embed_apikey_edit.setPlaceholderText("留空则复用上方 LLM 的 API Key")
        apikey_row = QHBoxLayout()
        apikey_row.addWidget(self._embed_apikey_label)
        apikey_row.addWidget(self._embed_apikey_edit)
        embed_layout.addLayout(apikey_row)

        self._embed_baseurl_label = QLabel("Base URL:")
        self._embed_baseurl_label.setFixedWidth(90)
        self._embed_baseurl_edit = QLineEdit()
        self._embed_baseurl_edit.setPlaceholderText("留空则复用上方 LLM 的 Base URL")
        baseurl_row = QHBoxLayout()
        baseurl_row.addWidget(self._embed_baseurl_label)
        baseurl_row.addWidget(self._embed_baseurl_edit)
        embed_layout.addLayout(baseurl_row)

        tab_llm_layout.addWidget(embed_box)
        tab_llm_layout.addStretch()
        self._tabs.addTab(tab_llm, "LLM 与模型")

        # ── 术语库配置区 ──────────────────────────────────────────────────────
        term_box = QGroupBox("术语库来源（上方优先级更高）")
        term_layout = QVBoxLayout(term_box)

        self._priority_list = QListWidget()
        self._priority_list.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        self._priority_list.setMaximumHeight(110)
        for source_name in ["dynamic（动态词库）", "paratranz（ParaTranz 术语）",
                             "json（本地 JSON）", "excel（本地 Excel）"]:
            self._priority_list.addItem(QListWidgetItem(source_name))
        term_layout.addWidget(self._priority_list)

        json_row = QHBoxLayout()
        json_row.addWidget(QLabel("本地 JSON:"))
        self._json_path_edit = QLineEdit()
        self._json_path_edit.setPlaceholderText("可选")
        json_row.addWidget(self._json_path_edit)
        json_browse = QPushButton("浏览")
        json_browse.setFixedWidth(52)
        json_browse.clicked.connect(lambda: self._browse_file(self._json_path_edit, "JSON 文件 (*.json)"))
        json_row.addWidget(json_browse)
        term_layout.addLayout(json_row)

        excel_row = QHBoxLayout()
        excel_row.addWidget(QLabel("本地 Excel:"))
        self._excel_path_edit = QLineEdit()
        self._excel_path_edit.setPlaceholderText("可选")
        excel_row.addWidget(self._excel_path_edit)
        excel_browse = QPushButton("浏览")
        excel_browse.setFixedWidth(52)
        excel_browse.clicked.connect(lambda: self._browse_file(self._excel_path_edit, "Excel 文件 (*.xlsx *.xls)"))
        excel_row.addWidget(excel_browse)
        term_layout.addLayout(excel_row)

        excel_col_row = QHBoxLayout()
        excel_col_row.addWidget(QLabel("原文列:"))
        self._excel_orig_col_edit = QLineEdit("A")
        self._excel_orig_col_edit.setFixedWidth(40)
        excel_col_row.addWidget(self._excel_orig_col_edit)
        excel_col_row.addWidget(QLabel("译文列:"))
        self._excel_trans_col_edit = QLineEdit("B")
        self._excel_trans_col_edit.setFixedWidth(40)
        excel_col_row.addWidget(self._excel_trans_col_edit)
        excel_col_row.addStretch()
        term_layout.addLayout(excel_col_row)

        view_terms_btn = QPushButton("查看/编辑动态术语库")
        view_terms_btn.clicked.connect(self._on_view_terms)
        term_layout.addWidget(view_terms_btn)

        # Tab 2: 术语库
        tab_terms = QWidget()
        tab_terms_layout = QVBoxLayout(tab_terms)
        tab_terms_layout.setSpacing(6)
        tab_terms_layout.addWidget(term_box)
        tab_terms_layout.addStretch()
        self._tabs.addTab(tab_terms, "术语库")

        # ── 翻译范围区 ────────────────────────────────────────────────────────
        scope_box = QGroupBox("翻译范围")
        scope_layout = QVBoxLayout(scope_box)
        scope_layout.setSpacing(4)

        # 作用域状态
        self._scope_stage_filters: set[int] = set()
        self._scope_label_filters: set[str] = set()
        self._scope_category_filters: set[str] = set()
        self._scope_preset: str | None = None

        # 快捷预设按钮
        preset_row = QHBoxLayout()
        preset_row.addWidget(QLabel("快捷："))
        self._preset_untranslated = QPushButton("全部未翻译")
        self._preset_untranslated.setCursor(Qt.CursorShape.PointingHandCursor)
        self._preset_untranslated.clicked.connect(lambda: self._on_preset("untranslated"))
        preset_row.addWidget(self._preset_untranslated)
        self._preset_table_view = QPushButton("当前主表视图")
        self._preset_table_view.setCursor(Qt.CursorShape.PointingHandCursor)
        self._preset_table_view.clicked.connect(lambda: self._on_preset("table_view"))
        preset_row.addWidget(self._preset_table_view)
        preset_row.addStretch()
        scope_layout.addLayout(preset_row)

        # 翻译状态维度标签
        stage_row = QHBoxLayout()
        stage_row.setSpacing(3)
        stage_row.addWidget(QLabel("状态："))
        self._scope_stage_all_btn = QPushButton("不限")
        self._scope_stage_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._scope_stage_all_btn.clicked.connect(lambda: self._on_scope_stage_clicked(None))
        stage_row.addWidget(self._scope_stage_all_btn)
        self._scope_stage_btns: dict[int, QPushButton] = {}
        stage_row.addStretch()
        scope_layout.addLayout(stage_row)

        # 标记维度标签
        mark_row = QHBoxLayout()
        mark_row.setSpacing(3)
        mark_row.addWidget(QLabel("标记："))
        self._scope_label_all_btn = QPushButton("不限")
        self._scope_label_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._scope_label_all_btn.clicked.connect(lambda: self._on_scope_label_clicked(None))
        mark_row.addWidget(self._scope_label_all_btn)
        self._scope_label_btns: dict[str, QPushButton] = {}
        mark_row.addStretch()
        scope_layout.addLayout(mark_row)

        # 分类维度标签
        cat_row = QHBoxLayout()
        cat_row.setSpacing(3)
        cat_row.addWidget(QLabel("分类："))
        self._scope_cat_all_btn = QPushButton("不限")
        self._scope_cat_all_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._scope_cat_all_btn.clicked.connect(lambda: self._on_scope_category_clicked(None))
        cat_row.addWidget(self._scope_cat_all_btn)
        self._scope_cat_btns: dict[str, QPushButton] = {}
        cat_row.addStretch()
        scope_layout.addLayout(cat_row)

        self._overwrite_check = QCheckBox("覆盖已有译文（重新翻译）")
        scope_layout.addWidget(self._overwrite_check)

        self._estimate_lbl = QLabel("预计：— 条")
        self._estimate_lbl.setStyleSheet("color: #888; font-size: 11px;")
        scope_layout.addWidget(self._estimate_lbl)

        # ── 混合模式面板 ──────────────────────────────────────────────────────
        mixed_panel = QWidget()
        mixed_layout = QVBoxLayout(mixed_panel)
        mixed_layout.setContentsMargins(0, 0, 0, 0)
        from ._rule_editor_widget import _RuleEditorWidget
        self._rule_editor = _RuleEditorWidget()
        mixed_layout.addWidget(self._rule_editor)
        order_row = QHBoxLayout()
        order_row.addWidget(QLabel("执行顺序:"))
        self._order_combo = QComboBox()
        self._order_combo.addItems(["串行（先翻译后润色）", "并行"])
        order_row.addWidget(self._order_combo)
        order_row.addStretch()
        mixed_layout.addLayout(order_row)
        mixed_estimate = QLabel("预计：— 条")
        mixed_estimate.setStyleSheet("color: #888; font-size: 11px;")
        self._mixed_estimate_lbl = mixed_estimate
        mixed_layout.addWidget(mixed_estimate)

        self._scope_stack = QStackedWidget()
        self._scope_stack.addWidget(scope_box)
        self._scope_stack.addWidget(mixed_panel)

        # ── 后处理配置区 ───────────────────────────────────────────────────────
        self._pp_box = QGroupBox("后处理配置")
        pp_layout = QVBoxLayout(self._pp_box)
        pp_layout.setSpacing(6)

        # 总开关
        self._pp_enable_check = QCheckBox("启用翻译后质量检查与优化")
        self._pp_enable_check.setChecked(True)
        self._pp_enable_check.setToolTip("启用后将对翻译结果进行质量检查、修复和润色，可能增加额外耗时和API调用")
        pp_layout.addWidget(self._pp_enable_check)

        # 分隔线
        line1 = QFrame()
        line1.setFrameShape(QFrame.Shape.HLine)
        line1.setStyleSheet("color: #ccc;")
        pp_layout.addWidget(line1)

        # 阶段1: 检测
        detect_label = QLabel("<b>阶段1: 质量检测</b>")
        pp_layout.addWidget(detect_label)

        self._pp_consistency_check = QCheckBox("术语一致性检查")
        self._pp_consistency_check.setChecked(True)
        self._pp_consistency_check.setToolTip("检查译文是否使用了术语表中的标准译法")
        pp_layout.addWidget(self._pp_consistency_check)

        self._pp_format_check = QCheckBox("格式验证（占位符、标签、引号等）")
        self._pp_format_check.setChecked(True)
        self._pp_format_check.setToolTip("检查译文是否保留了原文的占位符、格式标记和引号闭合")
        pp_layout.addWidget(self._pp_format_check)

        self._pp_quality_gate_check = QCheckBox("LLM质量检测")
        self._pp_quality_gate_check.setChecked(True)
        self._pp_quality_gate_check.setToolTip("使用LLM评估译文质量，识别漏翻、错翻等问题")
        pp_layout.addWidget(self._pp_quality_gate_check)

        # 分隔线
        line2 = QFrame()
        line2.setFrameShape(QFrame.Shape.HLine)
        line2.setStyleSheet("color: #ccc;")
        pp_layout.addWidget(line2)

        # 阶段2: 修复与润色
        refine_label = QLabel("<b>阶段2: 修复与润色</b>")
        pp_layout.addWidget(refine_label)

        self._pp_refinement_check = QCheckBox("启用LLM自动修复")
        self._pp_refinement_check.setChecked(True)
        self._pp_refinement_check.setToolTip("对检测出的问题使用LLM进行自动修复")
        pp_layout.addWidget(self._pp_refinement_check)

        self._pp_polish_check = QCheckBox("启用润色优化（需要额外LLM调用）")
        self._pp_polish_check.setChecked(False)
        self._pp_polish_check.setToolTip("对译文进行流畅度和风格优化，显著提升翻译质量但消耗更多API调用")
        pp_layout.addWidget(self._pp_polish_check)

        # 润色选项子布局
        polish_options = QHBoxLayout()
        polish_options.addSpacing(20)

        polish_options.addWidget(QLabel("润色范围:"))
        self._pp_polish_scope_combo = QComboBox()
        self._pp_polish_scope_combo.addItems(["全部条目", "仅通过检测的条目", "仅有问题需修复的条目"])
        self._pp_polish_scope_combo.setToolTip("全部: 润色所有译文\n仅通过: 只润色没有问题的译文\n仅问题: 只润色修复后的译文")
        polish_options.addWidget(self._pp_polish_scope_combo)

        polish_options.addSpacing(10)
        polish_options.addWidget(QLabel("润色强度:"))
        self._pp_polish_level_combo = QComboBox()
        self._pp_polish_level_combo.addItems(["轻微（仅修正明显错误）", "适中（平衡优化）", "深度（追求最佳表达）"])
        self._pp_polish_level_combo.setToolTip("轻微: 保守润色\n适中: 适度优化\n深度: 深度改写追求最佳表达")
        polish_options.addWidget(self._pp_polish_level_combo)

        polish_options.addStretch()
        pp_layout.addLayout(polish_options)

        # 润色预览确认
        self._polish_preview_check = QCheckBox("润色后预览确认（逐条对比接受/拒绝）")
        self._polish_preview_check.setChecked(False)
        self._polish_preview_check.setToolTip("勾选后，独立润色完成后弹出预览窗口，可逐条对比并选择接受或拒绝润色结果")
        pp_layout.addWidget(self._polish_preview_check)

        # 分隔线
        line3 = QFrame()
        line3.setFrameShape(QFrame.Shape.HLine)
        line3.setStyleSheet("color: #ccc;")
        pp_layout.addWidget(line3)

        # 阶段3: 裁决
        arbitrate_label = QLabel("<b>阶段3: 质量裁决</b>")
        pp_layout.addWidget(arbitrate_label)

        self._pp_arbitration_check = QCheckBox("启用LLM质量裁决")
        self._pp_arbitration_check.setChecked(True)
        self._pp_arbitration_check.setToolTip("对修复/润色后的译文进行最终质量裁决（通过/打回/待审）")
        pp_layout.addWidget(self._pp_arbitration_check)

        self._pp_strict_mode_check = QCheckBox("严格模式（质量存疑时直接打回而非标记待审）")
        self._pp_strict_mode_check.setChecked(False)
        self._pp_strict_mode_check.setToolTip("严格模式下，不确定质量的译文会被打回重翻而非保留待审")
        pp_layout.addWidget(self._pp_strict_mode_check)

        # 备注说明
        pp_note = QLabel("<i>提示：润色会在修复后执行，最终译文优先采用润色结果</i>")
        pp_note.setStyleSheet("color: #888; font-size: 11px;")
        pp_layout.addWidget(pp_note)

        # Tab 3: 后处理
        tab_pp = QWidget()
        tab_pp_layout = QVBoxLayout(tab_pp)
        tab_pp_layout.setSpacing(6)
        tab_pp_layout.addWidget(self._pp_box)
        tab_pp_layout.addStretch()
        self._tabs.addTab(tab_pp, "后处理")

        # ── 可滚动内容区 ──────────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(6)
        scroll_layout.addWidget(self._scope_stack)
        scroll_layout.addWidget(self._tabs)
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll, stretch=1)

        # ── 底部按钮 ──────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        self._history_btn = QPushButton("历史报告")
        self._history_btn.clicked.connect(self._on_open_history)
        btn_row.addWidget(self._history_btn)
        btn_row.addStretch()
        self._start_btn = QPushButton("▶ 开始翻译")
        self._start_btn.setStyleSheet(
            "QPushButton { background: #1976D2; color: white; font-weight: bold;"
            " padding: 6px 16px; border-radius: 4px; }"
            "QPushButton:hover { background: #1565C0; }"
            "QPushButton:disabled { background: #bbb; }"
        )
        self._start_btn.clicked.connect(self._on_start)
        btn_row.addWidget(self._start_btn)
        main_layout.addLayout(btn_row)

        self._mode_translate.toggled.connect(self._on_mode_changed)
        self._mode_polish.toggled.connect(self._on_mode_changed)
        self._mode_mixed.toggled.connect(self._on_mode_changed)
        self._overwrite_check.toggled.connect(self._update_estimate)

    # ── 断点检测 ──────────────────────────────────────────────────────────────

    def _check_checkpoint(self):
        esp_path = self._ctx.esp_path
        if not esp_path:
            return
        from src.transbridge.ai_translator.translator import ProgressCheckpoint
        cp = ProgressCheckpoint.load(esp_path)
        if cp is None:
            return
        done = len(cp.completed_fingerprints)
        reply = QMessageBox.question(
            self, "检测到未完成的翻译任务",
            f"上次翻译未完成（已完成 {done} 批，"
            f"成功 {cp.result_so_far.get('success_count', 0)} 条）。\n\n"
            "是否从断点继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes,
        )
        if reply == QMessageBox.StandardButton.No:
            cp.delete(esp_path)

    # ── 配置加载/保存 ─────────────────────────────────────────────────────────

    def _load_config(self):
        from src.transbridge.paratranz.config_manager import LLMConfig
        cfg = LLMConfig.load_from_file()
        self._provider_combo.setCurrentIndex(0 if cfg.provider != "anthropic" else 1)
        self._target_lang_combo.setCurrentText(cfg.target_lang)
        self._model_edit.setText(cfg.model)
        self._apikey_edit.setText(cfg.api_key)
        self._baseurl_edit.setText(cfg.base_url)
        self._concurrent_spin.setValue(cfg.max_concurrent)
        self._tokens_spin.setValue(cfg.max_tokens_per_batch)
        self._output_tokens_spin.setValue(cfg.max_output_tokens)
        self._max_terms_spin.setValue(cfg.max_terms_per_batch)
        self._json_path_edit.setText(cfg.local_json_path)
        self._excel_path_edit.setText(cfg.local_excel_path)
        self._excel_orig_col_edit.setText(cfg.excel_original_col)
        self._excel_trans_col_edit.setText(cfg.excel_translation_col)
        # 后处理配置
        self._pp_enable_check.setChecked(cfg.enable_post_process)
        self._pp_consistency_check.setChecked(cfg.pp_enable_consistency_check)
        self._pp_format_check.setChecked(cfg.pp_enable_format_validation)
        self._pp_quality_gate_check.setChecked(cfg.pp_enable_quality_gate)
        self._pp_refinement_check.setChecked(cfg.pp_enable_refinement)
        self._pp_polish_check.setChecked(cfg.pp_enable_polish)
        # 润色范围映射
        scope_map = {"all": 0, "passed": 1, "has_issues": 2}
        self._pp_polish_scope_combo.setCurrentIndex(scope_map.get(cfg.pp_polish_scope, 0))
        # 润色强度映射
        level_map = {"light": 0, "moderate": 1, "aggressive": 2}
        self._pp_polish_level_combo.setCurrentIndex(level_map.get(cfg.pp_polish_level, 1))
        self._pp_arbitration_check.setChecked(cfg.pp_enable_arbitration)
        self._pp_strict_mode_check.setChecked(cfg.pp_strict_arbitration)
        self._polish_preview_check.setChecked(cfg.polish_preview_enabled)
        # 更新控件状态
        self._on_pp_enable_changed()
        # Embedding 配置
        self._embed_provider_combo.setCurrentIndex(0 if cfg.embedding.provider == "local" else 1)
        self._embed_local_model_edit.setText(cfg.embedding.local_model_path)
        self._embed_model_edit.setText(cfg.embedding.model)
        self._embed_apikey_edit.setText(cfg.embedding.api_key)
        self._embed_baseurl_edit.setText(cfg.embedding.base_url)
        priority_map = {
            "dynamic": "dynamic（动态词库）",
            "paratranz": "paratranz（ParaTranz 术语）",
            "json": "json（本地 JSON）",
            "excel": "excel（本地 Excel）",
        }
        if cfg.term_priority:
            self._priority_list.clear()
            for p in cfg.term_priority:
                if p in priority_map:
                    self._priority_list.addItem(QListWidgetItem(priority_map[p]))
        self._on_provider_changed()
        self._on_embed_provider_changed()
        self._update_estimate()
        self._update_paratranz_item_style()

    def _update_paratranz_item_style(self):
        """若未选择 ParaTranz 项目，将列表中 paratranz 来源项显示为灰色并加提示。"""
        no_project = not self._ctx.current_project
        for i in range(self._priority_list.count()):
            item = self._priority_list.item(i)
            if "paratranz" in item.text():
                if no_project:
                    item.setForeground(QBrush(QColor("#999999")))
                    item.setToolTip("未选择 ParaTranz 项目，此来源将被跳过")
                else:
                    item.setForeground(QBrush(QColor()))  # 恢复默认颜色
                    item.setToolTip("")
                break

    def _save_config(self):
        from src.transbridge.paratranz.config_manager import LLMConfig
        cfg = LLMConfig.load_from_file()
        cfg.provider = "anthropic" if self._provider_combo.currentIndex() == 1 else "openai_compatible"
        cfg.target_lang = self._target_lang_combo.currentText()
        cfg.model = self._model_edit.text().strip()
        cfg.api_key = self._apikey_edit.text().strip()
        cfg.base_url = self._baseurl_edit.text().strip()
        cfg.max_concurrent = self._concurrent_spin.value()
        cfg.max_tokens_per_batch = self._tokens_spin.value()
        cfg.max_output_tokens = self._output_tokens_spin.value()
        cfg.max_terms_per_batch = self._max_terms_spin.value()
        cfg.local_json_path = self._json_path_edit.text().strip()
        cfg.local_excel_path = self._excel_path_edit.text().strip()
        cfg.excel_original_col = self._excel_orig_col_edit.text().strip() or "A"
        cfg.excel_translation_col = self._excel_trans_col_edit.text().strip() or "B"
        # 后处理配置
        cfg.enable_post_process = self._pp_enable_check.isChecked()
        cfg.pp_enable_consistency_check = self._pp_consistency_check.isChecked()
        cfg.pp_enable_format_validation = self._pp_format_check.isChecked()
        cfg.pp_enable_quality_gate = self._pp_quality_gate_check.isChecked()
        cfg.pp_enable_refinement = self._pp_refinement_check.isChecked()
        cfg.pp_enable_polish = self._pp_polish_check.isChecked()
        scope_map = {0: "all", 1: "passed", 2: "has_issues"}
        cfg.pp_polish_scope = scope_map.get(self._pp_polish_scope_combo.currentIndex(), "all")
        level_map = {0: "light", 1: "moderate", 2: "aggressive"}
        cfg.pp_polish_level = level_map.get(self._pp_polish_level_combo.currentIndex(), "moderate")
        cfg.pp_enable_arbitration = self._pp_arbitration_check.isChecked()
        cfg.pp_strict_arbitration = self._pp_strict_mode_check.isChecked()
        cfg.polish_preview_enabled = self._polish_preview_check.isChecked()
        # Embedding 配置
        cfg.embedding.provider = "local" if self._embed_provider_combo.currentIndex() == 0 else "openai"
        cfg.embedding.local_model_path = self._embed_local_model_edit.text().strip()
        cfg.embedding.model = self._embed_model_edit.text().strip()
        cfg.embedding.api_key = self._embed_apikey_edit.text().strip()
        cfg.embedding.base_url = self._embed_baseurl_edit.text().strip()
        key_map = {
            "dynamic（动态词库）": "dynamic",
            "paratranz（ParaTranz 术语）": "paratranz",
            "json（本地 JSON）": "json",
            "excel（本地 Excel）": "excel",
        }
        cfg.term_priority = [
            key_map[self._priority_list.item(i).text()]
            for i in range(self._priority_list.count())
            if self._priority_list.item(i).text() in key_map
        ]
        cfg.save_to_file()
        return cfg

    def _schedule_save(self):
        """防抖保存：延迟 2 秒，期间有新变更则重置计时器。"""
        if self._save_timer is None:
            from PyQt6.QtCore import QTimer
            self._save_timer = QTimer(self)
            self._save_timer.setSingleShot(True)
            self._save_timer.timeout.connect(self._save_config)
        self._save_timer.start(2000)

    def _connect_auto_save(self):
        """在配置加载完成后连接所有控件的变更信号，实现防抖自动保存。"""
        self._save_timer = None  # 延迟初始化，避免 _load_config 触发保存
        self._provider_combo.currentIndexChanged.connect(self._schedule_save)
        self._target_lang_combo.currentIndexChanged.connect(self._schedule_save)
        self._model_edit.textChanged.connect(self._schedule_save)
        self._apikey_edit.textChanged.connect(self._schedule_save)
        self._baseurl_edit.textChanged.connect(self._schedule_save)
        self._concurrent_spin.valueChanged.connect(self._schedule_save)
        self._tokens_spin.valueChanged.connect(self._schedule_save)
        self._output_tokens_spin.valueChanged.connect(self._schedule_save)
        self._max_terms_spin.valueChanged.connect(self._schedule_save)
        self._json_path_edit.textChanged.connect(self._schedule_save)
        self._excel_path_edit.textChanged.connect(self._schedule_save)
        self._excel_orig_col_edit.textChanged.connect(self._schedule_save)
        self._excel_trans_col_edit.textChanged.connect(self._schedule_save)
        self._priority_list.model().rowsMoved.connect(self._schedule_save)
        # 后处理配置自动保存
        self._pp_enable_check.toggled.connect(self._schedule_save)
        self._pp_consistency_check.toggled.connect(self._schedule_save)
        self._pp_format_check.toggled.connect(self._schedule_save)
        self._pp_quality_gate_check.toggled.connect(self._schedule_save)
        self._pp_refinement_check.toggled.connect(self._schedule_save)
        self._pp_polish_check.toggled.connect(self._schedule_save)
        self._pp_polish_scope_combo.currentIndexChanged.connect(self._schedule_save)
        self._pp_polish_level_combo.currentIndexChanged.connect(self._schedule_save)
        self._pp_arbitration_check.toggled.connect(self._schedule_save)
        self._pp_strict_mode_check.toggled.connect(self._schedule_save)
        self._polish_preview_check.toggled.connect(self._schedule_save)
        # 后处理控件联动
        self._pp_enable_check.toggled.connect(self._on_pp_enable_changed)
        self._pp_polish_check.toggled.connect(self._on_polish_changed)
        # Embedding 配置自动保存
        self._embed_provider_combo.currentIndexChanged.connect(self._schedule_save)
        self._embed_local_model_edit.textChanged.connect(self._schedule_save)
        self._embed_model_edit.textChanged.connect(self._schedule_save)
        self._embed_apikey_edit.textChanged.connect(self._schedule_save)
        self._embed_baseurl_edit.textChanged.connect(self._schedule_save)

    def _build_llm_config(self):
        from src.transbridge.paratranz.config_manager import LLMConfig
        cfg = LLMConfig()
        cfg.provider = "anthropic" if self._provider_combo.currentIndex() == 1 else "openai_compatible"
        cfg.target_lang = self._target_lang_combo.currentText()
        cfg.model = self._model_edit.text().strip()
        cfg.api_key = self._apikey_edit.text().strip()
        cfg.base_url = self._baseurl_edit.text().strip() or "https://api.openai.com/v1"
        cfg.max_concurrent = self._concurrent_spin.value()
        cfg.max_tokens_per_batch = self._tokens_spin.value()
        cfg.max_output_tokens = self._output_tokens_spin.value()
        cfg.max_terms_per_batch = self._max_terms_spin.value()
        cfg.local_json_path = self._json_path_edit.text().strip()
        cfg.local_excel_path = self._excel_path_edit.text().strip()
        cfg.excel_original_col = self._excel_orig_col_edit.text().strip() or "A"
        cfg.excel_translation_col = self._excel_trans_col_edit.text().strip() or "B"
        # Embedding 配置
        cfg.embedding.provider = "local" if self._embed_provider_combo.currentIndex() == 0 else "openai"
        cfg.embedding.local_model_path = self._embed_local_model_edit.text().strip()
        cfg.embedding.model = self._embed_model_edit.text().strip()
        cfg.embedding.api_key = self._embed_apikey_edit.text().strip()
        cfg.embedding.base_url = self._embed_baseurl_edit.text().strip()
        key_map = {
            "dynamic（动态词库）": "dynamic",
            "paratranz（ParaTranz 术语）": "paratranz",
            "json（本地 JSON）": "json",
            "excel（本地 Excel）": "excel",
        }
        cfg.term_priority = [
            key_map[self._priority_list.item(i).text()]
            for i in range(self._priority_list.count())
            if self._priority_list.item(i).text() in key_map
        ]
        return cfg

    # ── 事件处理 ──────────────────────────────────────────────────────────────

    def _on_provider_changed(self):
        is_openai = self._provider_combo.currentIndex() == 0
        self._baseurl_edit.setEnabled(is_openai)

    def _on_mode_changed(self):
        """模式切换时调整UI。"""
        is_polish = self._mode_polish.isChecked()
        is_mixed = self._mode_mixed.isChecked()
        # 显示/隐藏覆盖策略和面板
        self._overwrite_check.setVisible(not is_polish and not is_mixed)
        if is_mixed:
            self._scope_stack.setCurrentIndex(1)  # 混合面板
            self._start_btn.setText("▶ 开始执行")
        else:
            self._scope_stack.setCurrentIndex(0)  # 标准作用域面板
            self._reset_scope_to_default(is_polish)
            if is_polish:
                self._start_btn.setText("▶ 开始润色")
            else:
                self._start_btn.setText("▶ 开始翻译")
        self._update_estimate()

    # ── 作用域方法 ─────────────────────────────────────────────────────────

    def _reset_scope_to_default(self, is_polish: bool):
        self._scope_stage_filters.clear()
        self._scope_label_filters.clear()
        self._scope_category_filters.clear()
        self._scope_preset = None
        if is_polish:
            self._scope_stage_filters = {1, 2, 3, 5}
        else:
            self._scope_stage_filters = {0}
        self._rebuild_scope_tags()

    def _on_preset(self, preset: str):
        self._scope_stage_filters.clear()
        self._scope_label_filters.clear()
        self._scope_category_filters.clear()
        if preset == "untranslated":
            self._scope_stage_filters = {0}
            self._scope_preset = None
        elif preset == "table_view":
            self._scope_preset = "table_view"
        self._rebuild_scope_tags()
        self._update_estimate()

    def _on_scope_stage_clicked(self, stage: int | None):
        if stage is None:
            self._scope_stage_filters.clear()
        elif stage in self._scope_stage_filters:
            self._scope_stage_filters.discard(stage)
        else:
            self._scope_stage_filters.add(stage)
        self._scope_preset = None
        self._rebuild_scope_tags()
        self._update_estimate()

    def _on_scope_label_clicked(self, mark: str | None):
        if mark is None:
            self._scope_label_filters.clear()
        elif mark in self._scope_label_filters:
            self._scope_label_filters.discard(mark)
        else:
            self._scope_label_filters.add(mark)
        self._scope_preset = None
        self._rebuild_scope_tags()
        self._update_estimate()

    def _on_scope_category_clicked(self, cat: str | None):
        if cat is None:
            self._scope_category_filters.clear()
        elif cat in self._scope_category_filters:
            self._scope_category_filters.discard(cat)
        else:
            self._scope_category_filters.add(cat)
        self._scope_preset = None
        self._rebuild_scope_tags()
        self._update_estimate()

    def _rebuild_scope_tags(self):
        from collections import Counter
        from src.transbridge.converter.translation_entry import STAGE_LABELS
        from src.transbridge.ui.workbench.step2 import _ALL_CATEGORIES

        collection = self._ctx.collection
        entries = list(collection) if collection else []

        # Stage 标签
        counter = Counter()
        for e in entries:
            counter[e.stage] += 1
        for stage_val, label in STAGE_LABELS.items():
            if stage_val in self._scope_stage_btns:
                continue  # 已存在，更新样式
            btn = QPushButton(f"{label} {counter.get(stage_val, 0)}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked, s=stage_val: self._on_scope_stage_clicked(s))
            self._scope_stage_btns[stage_val] = btn
            # 找到 stage_row 并插入
            row_layout = self._scope_stage_all_btn.parent().layout()
            if row_layout:
                row_layout.insertWidget(row_layout.count() - 1, btn)  # stretch 之前
        for stage_val, btn in self._scope_stage_btns.items():
            count = counter.get(stage_val, 0)
            label = STAGE_LABELS.get(stage_val, "?")
            btn.setText(f"{label} {count}")
            active = stage_val in self._scope_stage_filters
            btn.setStyleSheet(
                "QPushButton { background: #2196F3; color: white; font-weight: bold; padding: 2px 8px; border-radius: 6px; }"
                if active else
                "QPushButton { background: #f0f0f0; border: 1px solid #ccc; padding: 2px 8px; border-radius: 6px; }"
            )

        # 标签维度（从主表标签库读取）
        label_library = self._step2._label_library if hasattr(self._step2, '_label_library') else {}
        entry_labels = self._step2._entry_labels if hasattr(self._step2, '_entry_labels') else {}
        label_counter = Counter()
        for labels in entry_labels.values():
            for lid in labels:
                label_counter[lid] += 1
        for lid, info in label_library.items():
            if lid in self._scope_label_btns:
                continue
            btn = QPushButton(f"● {info['name']} {label_counter.get(lid, 0)}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked, l=lid: self._on_scope_label_clicked(l))
            self._scope_label_btns[lid] = btn
            row_layout = self._scope_label_all_btn.parent().layout()
            if row_layout:
                row_layout.insertWidget(row_layout.count() - 1, btn)
        for lid, btn in self._scope_label_btns.items():
            count = label_counter.get(lid, 0)
            info = label_library.get(lid, {})
            label_name = info.get('name', '?')
            btn.setText(f"● {label_name} {count}")
            active = lid in self._scope_label_filters
            btn.setStyleSheet(
                "QPushButton { background: #2196F3; color: white; font-weight: bold; padding: 2px 8px; border-radius: 6px; }"
                if active else
                "QPushButton { background: #f0f0f0; border: 1px solid #ccc; padding: 2px 8px; border-radius: 6px; }"
            )

        # Category 标签
        cat_counter = Counter()
        from src.transbridge.ui.workbench.step2 import _entry_category
        for e in entries:
            cat_counter[_entry_category(e)] += 1
        for cat in _ALL_CATEGORIES:
            if cat in self._scope_cat_btns:
                continue
            btn = QPushButton(f"{cat} {cat_counter.get(cat, 0)}")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda checked, c=cat: self._on_scope_category_clicked(c))
            self._scope_cat_btns[cat] = btn
            row_layout = self._scope_cat_all_btn.parent().layout()
            if row_layout:
                row_layout.insertWidget(row_layout.count() - 1, btn)
        for cat, btn in self._scope_cat_btns.items():
            count = cat_counter.get(cat, 0)
            btn.setText(f"{cat} {count}")
            active = cat in self._scope_category_filters
            btn.setStyleSheet(
                "QPushButton { background: #2196F3; color: white; font-weight: bold; padding: 2px 8px; border-radius: 6px; }"
                if active else
                "QPushButton { background: #f0f0f0; border: 1px solid #ccc; padding: 2px 8px; border-radius: 6px; }"
            )

    def _build_scope_candidates(self) -> list:
        """按三维度筛选候选条目，与主表完全解耦。"""
        from src.transbridge.converter.translation_entry import STAGE_LOCKED, STAGE_HIDDEN
        from src.transbridge.ui.workbench.step2 import _entry_category

        collection = self._ctx.collection
        if collection is None:
            return []

        if self._scope_preset == "table_view":
            return self._step2._apply_all_filters()

        candidates = list(collection)

        if self._scope_stage_filters:
            candidates = [e for e in candidates if e.stage in self._scope_stage_filters]

        if self._scope_label_filters:
            entry_labels = self._step2._entry_labels if hasattr(self._step2, '_entry_labels') else {}
            candidates = [e for e in candidates
                          if e.id and entry_labels.get(e.id, set()) & self._scope_label_filters]

        if self._scope_category_filters:
            candidates = [e for e in candidates if _entry_category(e) in self._scope_category_filters]

        # 始终排除已锁定和已隐藏
        candidates = [e for e in candidates if e.stage not in (STAGE_LOCKED, STAGE_HIDDEN)]

        return candidates

    def _on_embed_provider_changed(self):
        """Embedding provider 切换时更新控件可见性。"""
        is_local = self._embed_provider_combo.currentIndex() == 0

        # 本地模型配置可见性
        self._embed_local_model_label.setVisible(is_local)
        self._embed_local_model_edit.setVisible(is_local)

        # API 配置可见性
        api_visible = not is_local
        self._embed_model_label.setVisible(api_visible)
        self._embed_model_edit.setVisible(api_visible)
        self._embed_apikey_label.setVisible(api_visible)
        self._embed_apikey_edit.setVisible(api_visible)
        self._embed_baseurl_label.setVisible(api_visible)
        self._embed_baseurl_edit.setVisible(api_visible)

    def _on_pp_enable_changed(self):
        """后处理总开关切换时更新所有子控件的启用状态。"""
        enabled = self._pp_enable_check.isChecked()
        # 阶段1: 检测
        self._pp_consistency_check.setEnabled(enabled)
        self._pp_format_check.setEnabled(enabled)
        self._pp_quality_gate_check.setEnabled(enabled)
        # 阶段2: 修复
        self._pp_refinement_check.setEnabled(enabled)
        # 阶段3: 裁决（依赖于总开关）
        self._pp_arbitration_check.setEnabled(enabled)
        # 润色及其子选项（依赖于总开关和润色开关）
        self._pp_polish_check.setEnabled(enabled)
        self._on_polish_changed()
        # 严格模式（依赖于总开关和裁决开关）
        self._pp_strict_mode_check.setEnabled(enabled and self._pp_arbitration_check.isChecked())

    def _on_polish_changed(self):
        """润色开关切换时更新润色选项的启用状态。"""
        enabled = self._pp_enable_check.isChecked() and self._pp_polish_check.isChecked()
        self._pp_polish_scope_combo.setEnabled(enabled)
        self._pp_polish_level_combo.setEnabled(enabled)

    def _update_estimate(self):
        collection = self._ctx.collection
        if collection is None:
            self._estimate_lbl.setText("预计：— 条（需先加载集合）")
            if hasattr(self, '_mixed_estimate_lbl'):
                self._mixed_estimate_lbl.setText("预计：— 条（需先加载集合）")
            return

        is_mixed = self._mode_mixed.isChecked()
        if is_mixed:
            rules = self._rule_editor.get_rules()
            entries = list(collection)
            actions = apply_rules(rules, entries)
            t_count = sum(1 for a in actions.values() if a == "translate")
            p_count = sum(1 for a in actions.values() if a == "polish")
            self._mixed_estimate_lbl.setText(
                f"预计：翻译 {t_count} 条 + 润色 {p_count} 条"
                + ("（两者均为0，请调整规则）" if t_count == 0 and p_count == 0 else "")
            )
            return

        is_polish = self._mode_polish.isChecked()
        if is_polish:
            candidates = self._build_scope_candidates()
            self._estimate_lbl.setText(
                f"润色范围：{len(candidates)} 条已翻译词条"
            )
            return

        candidates = self._build_scope_candidates()
        overwrite = self._overwrite_check.isChecked()
        if not overwrite:
            candidates = [e for e in candidates if not e.translation or e.stage == 0]

        if not candidates:
            self._estimate_lbl.setText("预计：0 条（无匹配条目，请调整作用域）")
            return

        from src.transbridge.ai_translator.batch_planner import BatchPlanner
        planner = BatchPlanner(max_tokens_per_batch=self._tokens_spin.value())
        plan = planner.plan(candidates)
        self._estimate_lbl.setText(
            f"预计：{plan.total_entries()} 条"
            f"（第一轮: {sum(len(b.entries) for b in plan.round1)}"
            f"  第二轮: {sum(len(b.entries) for b in plan.round2)}"
            f"  第三轮: {sum(len(b.entries) for b in plan.round3)}）"
        )

    def _on_test_connection(self):
        cfg = self._build_llm_config()
        if not cfg.api_key:
            QMessageBox.warning(self, "测试连接", "请先填写 API Key。")
            return
        if not cfg.model:
            QMessageBox.warning(self, "测试连接", "请先填写模型名。")
            return
        try:
            from src.transbridge.infra.llm_client import create_llm_client
            client = create_llm_client(cfg)
            reply = client.chat([{"role": "user", "content": "Say 'OK' in one word."}], max_tokens=10)
            QMessageBox.information(self, "测试连接", f"连接成功！模型回复：{reply}")
        except Exception as exc:
            QMessageBox.critical(self, "测试连接失败", str(exc))

    def _browse_file(self, target_edit: QLineEdit, file_filter: str):
        path, _ = QFileDialog.getOpenFileName(self, "选择文件", "", file_filter)
        if path:
            target_edit.setText(path)

    def _on_view_terms(self):
        esp_path = self._ctx.esp_path
        if not esp_path:
            QMessageBox.warning(self, "术语库", "尚未加载 ESP 文件。")
            return
        from src.transbridge.ai_translator.term_database import DynamicTermDatabase
        db = DynamicTermDatabase(esp_path)
        db.load()
        dlg = _TermEditorDialog(db, parent=self)
        dlg.exec()

    def _on_start(self):
        if self._mode_mixed.isChecked():
            self._on_mixed_start()
            return
        if self._mode_polish.isChecked():
            self._on_polish_start()
            return

        collection = self._ctx.collection
        if not collection:
            QMessageBox.warning(self, "翻译", "请先加载词条集合。")
            return

        cfg = self._save_config()

        if not cfg.api_key:
            QMessageBox.warning(self, "翻译", "请先填写 API Key。")
            return
        if not cfg.model:
            QMessageBox.warning(self, "翻译", "请先填写模型名。")
            return
        if not self._ctx.esp_path:
            QMessageBox.warning(self, "翻译", "找不到 ESP 路径，请重新加载集合。")
            return

        if not self._ctx.current_project and self._check_all_terms_empty(cfg):
            reply = QMessageBox.question(
                self, "术语库为空",
                "当前未选择 ParaTranz 项目，且所有术语来源均为空。\n\n"
                "没有术语库辅助，翻译质量可能下降。是否继续？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # 确定翻译目标
        candidates = self._build_scope_candidates()
        if not candidates:
            QMessageBox.warning(self, "翻译", "当前作用域无匹配词条。")
            return
        target_ids = [e.id for e in candidates]

        # 加载断点（如有）
        from src.transbridge.ai_translator.translator import AutoTranslator, TranslatorConfig, ProgressCheckpoint
        checkpoint = ProgressCheckpoint.load(self._ctx.esp_path)

        translator_cfg = TranslatorConfig(
            llm_config=cfg,
            esp_path=self._ctx.esp_path,
            overwrite=self._overwrite_check.isChecked(),
        )

        paratranz_client = None
        project_id = None
        if self._ctx.current_project:
            project_id = self._ctx.current_project.get("id")
            from src.transbridge.paratranz.api.paratranz_terms_api import ParatranzTermsAPI
            paratranz_client = ParatranzTermsAPI(self._ctx.config)

        translator = AutoTranslator(translator_cfg, paratranz_client, project_id)
        worker = _TranslationWorker(translator, collection, target_ids, checkpoint)

        progress_win = _TranslationProgressWindow(worker, self._ctx)
        self.progress_window_created.emit(progress_win)

        progress_win.show()
        worker.start()
        self.close()

    def _on_mixed_start(self):
        """混合模式：规则匹配 → 拆分为翻译/润色条目 → MixedWorker（占位，S12实现）。"""
        collection = self._ctx.collection
        if not collection:
            QMessageBox.warning(self, "混合模式", "请先加载词条集合。")
            return

        rules = self._rule_editor.get_rules()
        entries = list(collection)
        actions = apply_rules(rules, entries)
        translate_entries = [e for e in entries if actions.get(e.id) == "translate"]
        polish_entries = [e for e in entries if actions.get(e.id) == "polish"]

        if not translate_entries and not polish_entries:
            QMessageBox.warning(self, "混合模式", "当前筛选条件下无匹配条目，请调整规则。")
            return

        cfg = self._save_config()
        # 保存执行顺序
        cfg.mixed_execution_order = "serial" if self._order_combo.currentIndex() == 0 else "parallel"
        # 保存规则
        cfg.action_rules = rules

        # 创建 _MixedWorker 并启动
        from ._mixed_worker import _MixedWorker
        self._mixed_worker = _MixedWorker(
            cfg=cfg,
            translate_entries=translate_entries,
            polish_entries=polish_entries,
            execution_order=cfg.mixed_execution_order,
            ctx=self._ctx,
        )
        self._mixed_worker.finished.connect(self._on_mixed_finished)
        self._mixed_worker.error.connect(lambda msg: QMessageBox.warning(self, "混合模式错误", msg))
        self._mixed_worker.start()
        QMessageBox.information(
            self, "混合模式",
            f"规则匹配完成：\n翻译 {len(translate_entries)} 条\n润色 {len(polish_entries)} 条\n\n"
            f"执行顺序：{'串行' if cfg.mixed_execution_order == 'serial' else '并行'}\n\n"
            f"混合执行已启动，完成后将弹出报告。",
        )

    def _on_mixed_finished(self, result: dict):
        """混合执行完成回调：汇总并弹出报告。"""
        t = result.get("translate")
        p = result.get("polish")
        lines = ["混合执行完成:"]
        if t:
            lines.append(f"翻译: 成功 {t.success_count}, 失败 {t.failed_count}")
        if p:
            lines.append(f"润色: 成功 {p.success_count}, 失败 {p.failed_count}")
            if hasattr(p, 'details') and p.details:
                failed = [d for d in p.details if not d['success']]
                if failed:
                    lines.append(f"润色失败条目 ({len(failed)}):")
                    for d in failed[:5]:
                        lines.append(f"  - {d['key']}: {d.get('error', '未知错误')[:50]}")
        QMessageBox.information(self, "混合模式", "\n".join(lines))

    def _on_polish_start(self):
        """润色模式：选中已翻译词条 → LLMPolisher → 可选预览 → 写入。"""
        collection = self._ctx.collection
        if not collection:
            QMessageBox.warning(self, "润色", "请先加载词条集合。")
            return

        cfg = self._save_config()

        if not cfg.api_key:
            QMessageBox.warning(self, "润色", "请先填写 API Key。")
            return
        if not cfg.model:
            QMessageBox.warning(self, "润色", "请先填写模型名。")
            return

        # 按作用域获取有译文的条目
        candidates = self._build_scope_candidates()
        entries_with_translation = [e for e in candidates if e.translation]
        if not entries_with_translation:
            QMessageBox.warning(self, "润色", "作用域内条目均无译文，无法润色。")
            return

        # 创建 LLM 客户端
        from src.transbridge.infra.llm_client import create_llm_client
        llm_client = create_llm_client(cfg)

        # 创建术语管理器（可选）
        term_manager = None
        if self._ctx.esp_path:
            from src.transbridge.ai_translator.term_database import DynamicTermDatabase, TermDatabaseManager
            dynamic_db = DynamicTermDatabase(self._ctx.esp_path)
            dynamic_db.load()
            term_manager = TermDatabaseManager([dynamic_db.as_list()])

        # 创建润色器
        from src.transbridge.ai_translator.post_processor.polisher import LLMPolisher
        polish_level = cfg.pp_polish_level or "moderate"
        polisher = LLMPolisher(
            llm_client=llm_client,
            term_manager=term_manager,
            game_profile=cfg.game_profile,
            target_lang=cfg.target_lang,
            polish_level=polish_level,
        )

        # 创建 Worker
        from src.transbridge.ui.tools.ai_translator._polish_worker import _PolishWorker
        worker = _PolishWorker(polisher, entries_with_translation)

        # 进度窗口（复用翻译进度窗口模式 — 简单弹窗）
        preview_enabled = self._polish_preview_check.isChecked()

        if preview_enabled:
            # 预览模式：先收集结果，再弹预览
            self._polish_with_preview(worker, entries_with_translation, collection)
        else:
            # 直接写入模式
            self._polish_direct(worker, entries_with_translation, collection)

    def _polish_direct(self, worker, entries, collection):
        """润色直接写入模式。"""
        progress_dlg = QDialog(self)
        progress_dlg.setWindowTitle("AI 润色 — 进行中")
        progress_dlg.resize(400, 100)
        progress_layout = QVBoxLayout(progress_dlg)
        progress_bar = QProgressBar()
        progress_bar.setRange(0, len(entries))
        progress_layout.addWidget(progress_bar)
        status_lbl = QLabel("准备中…")
        progress_layout.addWidget(status_lbl)

        worker.progress.connect(lambda c, t, m: (
            progress_bar.setValue(c),
            progress_bar.setMaximum(t),
            status_lbl.setText(m),
        ))

        worker.finished_all.connect(lambda results: self._on_polish_finished_direct(
            results, entries, collection, progress_dlg
        ))
        worker.error.connect(lambda err: (
            progress_dlg.close(),
            QMessageBox.critical(self, "润色错误", err),
        ))

        worker.start()
        progress_dlg.exec()

    def _polish_with_preview(self, worker, entries, collection):
        """润色预览模式。"""
        progress_dlg = QDialog(self)
        progress_dlg.setWindowTitle("AI 润色 — 进行中")
        progress_dlg.resize(400, 100)
        progress_layout = QVBoxLayout(progress_dlg)
        progress_bar = QProgressBar()
        progress_bar.setRange(0, len(entries))
        progress_layout.addWidget(progress_bar)
        status_lbl = QLabel("准备中…")
        progress_layout.addWidget(status_lbl)

        worker.progress.connect(lambda c, t, m: (
            progress_bar.setValue(c),
            progress_bar.setMaximum(t),
            status_lbl.setText(m),
        ))

        def _on_done(results):
            progress_dlg.close()
            self._last_polish_results = results
            from src.transbridge.ui.tools.ai_translator._polish_preview_dialog import _PolishPreviewDialog
            preview = _PolishPreviewDialog(entries, results, parent=self)
            if preview.exec() == QDialog.DialogCode.Accepted:
                self._apply_polish_results(entries, preview.get_results(), collection)
            worker.deleteLater()

        worker.finished_all.connect(_on_done)
        worker.error.connect(lambda err: (
            progress_dlg.close(),
            QMessageBox.critical(self, "润色错误", err),
        ))

        worker.start()
        progress_dlg.exec()

    def _on_polish_finished_direct(self, results, entries, collection, progress_dlg):
        """直接写入模式完成回调。"""
        progress_dlg.close()
        applied = 0
        failed = 0
        for entry in entries:
            result = results.get(entry.id)
            if result and result.polished_translation and result.confidence > 0:
                updated = type(entry)(
                    id=entry.id, key=entry.key, original=entry.original,
                    translation=result.polished_translation, stage=entry.stage,
                    context=entry.context, form_id_with_plugin=entry.form_id_with_plugin,
                    string_id=entry.string_id, dsd_type=entry.dsd_type,
                    dsd_index=entry.dsd_index, editor_id=entry.editor_id,
                )
                collection.add(updated, overwrite=True)
                applied += 1
            else:
                failed += 1
        self._ctx.collection_changed.emit(collection)
        self._show_polish_report(results, entries, applied, 0, failed)
        self.close()

    def _apply_polish_results(self, entries, polish_decisions, collection):
        """应用润色预览结果到集合。"""
        applied = 0
        rejected = 0
        for entry in entries:
            decision = polish_decisions.get(entry.id)
            if decision is not None:  # accepted polish
                updated = type(entry)(
                    id=entry.id, key=entry.key, original=entry.original,
                    translation=decision, stage=entry.stage,
                    context=entry.context, form_id_with_plugin=entry.form_id_with_plugin,
                    string_id=entry.string_id, dsd_type=entry.dsd_type,
                    dsd_index=entry.dsd_index, editor_id=entry.editor_id,
                )
                collection.add(updated, overwrite=True)
                applied += 1
            elif decision is None and entry.id in polish_decisions:
                rejected += 1
        self._ctx.collection_changed.emit(collection)
        self._show_polish_report(
            self._last_polish_results if hasattr(self, '_last_polish_results') else {},
            entries, applied, rejected, 0,
        )
        self.close()

    def _show_polish_report(self, results, entries, accepted, rejected, failed):
        """生成润色报告并弹出报告对话框。"""
        # 计算失败数（置信度为0的结果）
        if not failed:
            failed = sum(1 for r in results.values() if r.confidence == 0.0)
        avg_conf = (sum(r.confidence for r in results.values()) / len(results)) if results else 0
        stats = {
            "total": len(results),
            "accepted": accepted,
            "rejected": rejected,
            "failed": failed,
            "polish_level": getattr(self, '_cfg', None) and getattr(self._cfg, 'polish_level', 'moderate') or 'moderate',
            "avg_confidence": avg_conf,
        }

        esp_stem = "unknown"
        if self._ctx.esp_path:
            from pathlib import Path
            esp_stem = Path(self._ctx.esp_path).stem

        report_path = None
        try:
            from src.transbridge.ai_translator.post_processor.report_generator import ReportGenerator
            gen = ReportGenerator(esp_stem)
            report_path = gen.generate_polish_report(results, entries, stats)
        except Exception:
            pass

        from ._translation_report_dialog import _TranslationReportDialog
        dialog = _TranslationReportDialog(
            polish_entries=entries,
            polish_results_dict=results,
            polish_stats=stats,
            report_path=report_path,
        )
        main_win = self._find_main_window()
        if main_win and hasattr(main_win, '_on_report_entry_activated'):
            dialog.entry_activated.connect(main_win._on_report_entry_activated)
        dialog.show()

    def _on_open_history(self):
        """打开历史报告查看对话框。"""
        from ._report_history_dialog import _ReportHistoryDialog
        dialog = _ReportHistoryDialog(parent=self)
        dialog.show()

    @staticmethod
    def _find_main_window():
        from src.transbridge.ui.main_window import MainWindow
        from PyQt6.QtWidgets import QWidget as QW
        for widget in QW.topLevelWidgets():
            if isinstance(widget, MainWindow):
                return widget
        return None

    def _check_all_terms_empty(self, cfg) -> bool:
        """检查所有术语来源是否均为空。"""
        import os
        from src.transbridge.ai_translator.term_database import DynamicTermDatabase
        dynamic_db = DynamicTermDatabase(self._ctx.esp_path)
        dynamic_db.load()
        if dynamic_db.as_list():
            return False

        if cfg.local_json_path and os.path.exists(cfg.local_json_path):
            try:
                import json
                with open(cfg.local_json_path, encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list) and data:
                    return False
                if isinstance(data, dict) and data:
                    return False
            except Exception:
                pass

        if cfg.local_excel_path and os.path.exists(cfg.local_excel_path):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(cfg.local_excel_path, read_only=True, data_only=True)
                ws = wb.active
                col_orig = self._col_letter_to_index(cfg.excel_original_col or "A")
                col_trans = self._col_letter_to_index(cfg.excel_translation_col or "B")
                for row in ws.iter_rows(min_row=2, values_only=True):
                    try:
                        term = row[col_orig] if col_orig < len(row) else None
                        trans = row[col_trans] if col_trans < len(row) else None
                        if term and trans:
                            return False
                    except (IndexError, TypeError):
                        continue
            except Exception:
                pass

        return True

    @staticmethod
    def _col_letter_to_index(letter: str) -> int:
        """将列字母（A/B/AA 等）转换为 0 起始的列索引。"""
        letter = letter.upper().strip()
        idx = 0
        for ch in letter:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx - 1

    def _get_filtered_entry_ids(self) -> list[str] | None:
        from src.transbridge.ui.workbench.step2 import _COL_KEY
        result = []
        table = self._step2._table
        for row in range(table.rowCount()):
            if not table.isRowHidden(row):
                item = table.item(row, _COL_KEY)
                if item:
                    from src.transbridge.converter.translation_entry import TranslationEntry
                    e = item.data(Qt.ItemDataRole.UserRole)
                    if isinstance(e, TranslationEntry):
                        result.append(e.id)
        return result if result else None